"""把各模型的 Fama-MacBeth 与投资组合排序结果整理成学术表格。

默认输入：

    D_analysis/output/fund_consistency/*/fama_macbeth_results.csv

默认输出：

    D_analysis/output/fama_macbeth_results汇总.xlsx
    D_analysis/output/portfolio_sorting汇总.xlsx

工作簿按 ``I_visualization/model_map.html`` 中的模型区块分组，每组生成两张表：

1. ``*_模型比较``：每个“模型目录 × 模型参数”一行，展示样本月份、月均基金数、
   平均 R² 和平均调整 R²；
2. ``*_回归系数显著性``：每个“模型目录 × 变量主名称”一行、模型参数为列，
   交叉格中第一行是保留两位小数并附显著性星号的系数，第二行是括号内
   Newey-West t 值。变量名中已由列标题表达的参数后缀会被去掉，使同一解释
   变量在不同参数下的结果对齐到同一行。

模型地图没有收录、但磁盘上确实存在的旧结果会放入“未映射模型”的两张表，
避免静默漏掉文件。正式表不再展示 p 值和标准误，但仍用原始 ``p_value`` 生成
星号：***、**、* 分别对应 1%、5%、10% 显著性水平。

投资组合排序汇总会从每个模型的 ``portfolio_sorting`` 目录中选择编号最大的
``ps_*_summary.csv``。工作簿按模型地图顺序为每个模型单独生成一张 Sheet，
并把 long-short、t-stat 与 p-value 三行合并为“多空收益+星号 / 括号内 t 值”。

运行方式：

    .venv/bin/python D_analysis/scripts/result_sorting.py
"""

from __future__ import annotations

import argparse
import html
import re
from collections import OrderedDict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.theme import theme_xml


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT_ROOT = PROJECT_ROOT / "D_analysis" / "output" / "fund_consistency"
DEFAULT_MODEL_MAP = PROJECT_ROOT / "I_visualization" / "model_map.html"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "D_analysis" / "output" / "fama_macbeth_results汇总.xlsx"
DEFAULT_PORTFOLIO_OUTPUT_PATH = (
    PROJECT_ROOT / "D_analysis" / "output" / "portfolio_sorting汇总.xlsx"
)

RESULT_FILE_NAME = "fama_macbeth_results.csv"
PORTFOLIO_SUMMARY_PATTERN = re.compile(r"ps_(\d+)_summary\.csv$")
UNMAPPED_SHEET_TITLE = "未映射模型"
FAMILY_SHEET_NAMES = {
    "fam-1": "1_FAC核心基础",
    "fam-2": "2_排名均值三分组",
    "fam-3": "3_跨期限替代指标",
    "fam-4": "4_热力图窗口搜索",
    "fam-5": "5_Winrate_HitRate",
    "fam-6": "6_市态条件一致性",
}

REQUIRED_COLUMNS = (
    "factor",
    "variable",
    "coef",
    "t_stat",
    "p_value",
    "n_months",
    "avg_monthly_n",
    "avg_r_squared",
    "avg_adj_r_squared",
)

COMPARISON_HEADERS = (
    "模型目录",
    "模型参数",
    "月份数",
    "月均基金数",
    "平均 R²",
    "平均调整 R²",
)

LATIN_FONT = "Times New Roman"
CHINESE_FONT = "微软雅黑"

PARAMETER_PATTERN = re.compile(r"m\d+_n\d+_pairwise\d+")
HEATMAP_N1_PARAMETER_PATTERN = re.compile(r"(?:^|_)m\d+_n1_pairwise\d+(?:_|$)")
HORIZON_PATTERN = re.compile(r"(\d+m(?:_\d+m)+)$")
CHINESE_SCRIPT_PATTERN = re.compile(
    r"[\u3000-\u303f\u3400-\u4dbf\u4e00-\u9fff\uff00-\uffef]"
)


def parse_args() -> argparse.Namespace:
    """读取命令行参数，方便在测试或其他项目副本中复用脚本。"""
    parser = argparse.ArgumentParser(
        description="汇总并格式化 Fama-MacBeth 与投资组合排序结果。"
    )
    parser.add_argument(
        "--input-root",
        type=Path,
        default=DEFAULT_INPUT_ROOT,
        help=f"模型结果根目录，默认：{DEFAULT_INPUT_ROOT}",
    )
    parser.add_argument(
        "--model-map",
        type=Path,
        default=DEFAULT_MODEL_MAP,
        help=f"模型地图 HTML，默认：{DEFAULT_MODEL_MAP}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"输出 Excel 路径，默认：{DEFAULT_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--portfolio-output",
        type=Path,
        default=DEFAULT_PORTFOLIO_OUTPUT_PATH,
        help=f"投资组合排序汇总路径，默认：{DEFAULT_PORTFOLIO_OUTPUT_PATH}",
    )
    return parser.parse_args()


def clean_model_key(raw_key: str) -> str:
    """去掉页面上的 ``（M0）`` 等展示标签，得到真实的模型目录名。"""
    return html.unescape(raw_key).strip().split("（", maxsplit=1)[0]


def clean_family_title(raw_title: str) -> str:
    """把 HTML 区块标题压缩成合法且易读的 Excel Sheet 名。"""
    title = html.unescape(raw_title).strip()
    # Sheet 名最多 31 个字符，且不能包含下列特殊字符。
    title = re.sub(r"[\\/*?:\[\]]", "_", title)
    title = re.sub(r"\s+", "", title)
    return title[:31] or "未命名分组"


def parse_model_groups(model_map_path: Path) -> OrderedDict[str, list[str]]:
    """按模型地图中的 ``section.family`` 顺序解析正式模型分组。

    这里只读取 id 为 ``fam-数字`` 的六个 registry 区块，主动排除页面后面的
    “registry 之外”补充内容。解析结果保留 HTML 中的顺序，用于稳定控制 Sheet
    顺序和 Sheet 内模型顺序。
    """
    if not model_map_path.exists():
        raise FileNotFoundError(f"模型地图不存在：{model_map_path}")

    text = model_map_path.read_text(encoding="utf-8")
    section_pattern = re.compile(
        r'<section\s+class="family"\s+id="(fam-\d+)"\s+data-family="([^"]+)">(.*?)</section>',
        flags=re.DOTALL,
    )
    key_pattern = re.compile(
        r'<td\s+class="col-key"><code>(fm_[^<]+)</code>',
        flags=re.DOTALL,
    )

    groups: OrderedDict[str, list[str]] = OrderedDict()
    seen_models: set[str] = set()
    for family_id, raw_title, section_html in section_pattern.findall(text):
        sheet_title = FAMILY_SHEET_NAMES.get(family_id, clean_family_title(raw_title))
        model_keys = [clean_model_key(item) for item in key_pattern.findall(section_html)]
        if not model_keys:
            continue

        duplicated = seen_models.intersection(model_keys)
        if duplicated:
            raise ValueError(f"模型地图存在跨区块重复模型：{sorted(duplicated)}")
        seen_models.update(model_keys)

        # 截断 Sheet 名后理论上可能重名；若发生则加稳定的数字后缀。
        base_title = sheet_title
        suffix = 2
        while sheet_title in groups:
            suffix_text = f"_{suffix}"
            sheet_title = f"{base_title[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        groups[sheet_title] = model_keys

    if not groups:
        raise ValueError(f"未能从模型地图解析任何 fam-* 模型区块：{model_map_path}")
    return groups


def significance_stars(p_value: object) -> str:
    """按照回归脚本相同的阈值返回显著性星号。"""
    value = pd.to_numeric(pd.Series([p_value]), errors="coerce").iloc[0]
    if pd.isna(value):
        return ""
    if value < 0.01:
        return "***"
    if value < 0.05:
        return "**"
    if value < 0.10:
        return "*"
    return ""


def format_two_decimals(value: object) -> str:
    """把回归统计量格式化为两位小数，同时避免出现难看的 ``-0.00``。"""
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return ""
    numeric = float(numeric)
    if abs(numeric) < 0.005:
        numeric = 0.0
    return f"{numeric:.2f}"


def format_estimate(coef: object, t_stat: object, p_value: object) -> str:
    """生成“系数+星号 / 括号内 t 值”的两行学术表格单元格。"""
    coef_text = format_two_decimals(coef)
    if not coef_text:
        return ""
    t_text = format_two_decimals(t_stat)
    stars = significance_stars(p_value)
    return f"{coef_text}{stars}\n({t_text})" if t_text else f"{coef_text}{stars}"


def format_portfolio_long_short(
    long_short_return: object,
    t_stat: object,
    p_value: object,
) -> str:
    """把多空收益格式化为“百分比+星号 / 括号内 t 值”的两行单元格。"""
    numeric_return = pd.to_numeric(
        pd.Series([long_short_return]), errors="coerce"
    ).iloc[0]
    if pd.isna(numeric_return):
        return ""
    return_text = format_two_decimals(float(numeric_return) * 100)
    t_text = format_two_decimals(t_stat)
    stars = significance_stars(p_value)
    return f"{return_text}%{stars}\n({t_text})" if t_text else f"{return_text}%{stars}"


def numeric_or_none(value: object) -> float | None:
    """把 CSV 字段转换为 Excel 数值；缺失值写成真正的空单元格。"""
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return None if pd.isna(numeric) else float(numeric)


def is_chinese_script_character(character: str) -> bool:
    """判断字符是否应使用中文字体，包含汉字和常见全角中文标点。"""
    return bool(CHINESE_SCRIPT_PATTERN.fullmatch(character))


def set_bilingual_text(
    cell,
    value: object,
    *,
    size: float,
    bold: bool = False,
    italic: bool = False,
    color: str = "000000",
) -> None:
    """把单元格中的中英文字符段分别写成微软雅黑和 Times New Roman。

    之前只依赖主题字体，部分 Excel 环境会继续显示本机默认字体。这次对纯中文、
    纯英文和中英混排分别写入明确的字体信息；混排文本使用富文本字符段，因此
    同一个单元格里的中文和英文也不会互相抢字体。
    """
    if value is None or not isinstance(value, str):
        cell.font = Font(name=LATIN_FONT, size=size, bold=bold, italic=italic, color=color)
        return

    if value == "":
        cell.font = Font(name=LATIN_FONT, size=size, bold=bold, italic=italic, color=color)
        return

    flags = [is_chinese_script_character(character) for character in value]
    if all(flags):
        cell.font = Font(name=CHINESE_FONT, size=size, bold=bold, italic=italic, color=color)
        return
    if not any(flags):
        cell.font = Font(name=LATIN_FONT, size=size, bold=bold, italic=italic, color=color)
        return

    # 中英混排时按连续字符段生成富文本，保证每一段都显式带有正确字体。
    blocks: list[TextBlock] = []
    start = 0
    current_flag = flags[0]
    for index in range(1, len(value) + 1):
        boundary = index == len(value) or flags[index] != current_flag
        if not boundary:
            continue
        text = value[start:index]
        blocks.append(
            TextBlock(
                InlineFont(
                    rFont=CHINESE_FONT if current_flag else LATIN_FONT,
                    sz=size,
                    b=bold,
                    i=italic,
                    color=color,
                ),
                text,
            )
        )
        if index < len(value):
            start = index
            current_flag = flags[index]

    cell.value = CellRichText(blocks)
    # 富文本已有逐段字体；基础字体作为不支持富文本查看器的回退。
    cell.font = Font(name=LATIN_FONT, size=size, bold=bold, italic=italic, color=color)


def build_bilingual_theme() -> bytes:
    """生成中英文分别使用指定字体的 Excel 主题。

    Excel 的主题字体可以按文字脚本选择字体，因此同一个单元格中的拉丁字符和
    简体中文也能分别显示为 Times New Roman 和微软雅黑。相比逐格判断“是否含
    中文”，这个方法不会把混合文本里的英文误设为中文字体。
    """
    custom_theme = theme_xml
    for font_group in ("majorFont", "minorFont"):
        group_pattern = re.compile(
            rf"(<a:{font_group}>)(.*?)(</a:{font_group}>)",
            flags=re.DOTALL,
        )
        match = group_pattern.search(custom_theme)
        if match is None:
            raise ValueError(f"默认 Excel 主题缺少 {font_group}，无法设置双语字体")

        group_body = match.group(2)
        group_body = re.sub(
            r'<a:latin typeface="[^"]*"/>',
            f'<a:latin typeface="{LATIN_FONT}"/>',
            group_body,
            count=1,
        )
        group_body = re.sub(
            r'<a:ea typeface="[^"]*"/>',
            f'<a:ea typeface="{CHINESE_FONT}"/>',
            group_body,
            count=1,
        )
        group_body = re.sub(
            r'<a:font script="Hans" typeface="[^"]*"/>',
            f'<a:font script="Hans" typeface="{CHINESE_FONT}"/>',
            group_body,
            count=1,
        )
        replacement = f"{match.group(1)}{group_body}{match.group(3)}"
        custom_theme = (
            custom_theme[: match.start()] + replacement + custom_theme[match.end() :]
        )

    return custom_theme.encode("utf-8")


def discover_result_files(input_root: Path) -> dict[str, Path]:
    """发现所有直接模型子目录中的结果文件，并以母文件夹名作为模型名。"""
    if not input_root.exists():
        raise FileNotFoundError(f"回归结果根目录不存在：{input_root}")

    result_files: dict[str, Path] = {}
    for path in sorted(input_root.glob(f"*/{RESULT_FILE_NAME}")):
        model_name = path.parent.name
        if model_name in result_files:
            raise ValueError(f"发现重复模型目录名：{model_name}")
        result_files[model_name] = path

    if not result_files:
        raise FileNotFoundError(
            f"在 {input_root} 下没有找到 */{RESULT_FILE_NAME}"
        )
    return result_files


def discover_latest_portfolio_summaries(input_root: Path) -> dict[str, Path]:
    """为每个模型选择编号最大的 ``ps_*_summary.csv``。

    ``ps_005`` 代表晚于 ``ps_004`` 的流水线结果，因此优先按编号判断“最新”；
    理论上同一编号只应有一个文件，若出现重复路径，则再以修改时间作为兜底。
    """
    if not input_root.exists():
        raise FileNotFoundError(f"回归结果根目录不存在：{input_root}")

    candidates: dict[str, list[tuple[int, int, Path]]] = {}
    for path in input_root.glob("*/portfolio_sorting/ps_*_summary.csv"):
        match = PORTFOLIO_SUMMARY_PATTERN.fullmatch(path.name)
        if match is None:
            continue
        model_name = path.parents[1].name
        candidates.setdefault(model_name, []).append(
            (int(match.group(1)), path.stat().st_mtime_ns, path)
        )

    if not candidates:
        raise FileNotFoundError(
            f"在 {input_root} 下没有找到 */portfolio_sorting/ps_*_summary.csv"
        )

    return {
        model_name: max(items, key=lambda item: (item[0], item[1]))[2]
        for model_name, items in candidates.items()
    }


def load_result_frame(model_name: str, csv_path: Path) -> pd.DataFrame:
    """读取单个模型结果，并补充模型目录与精简后的唯一参数名。"""
    frame = pd.read_csv(csv_path)
    missing = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"{csv_path} 缺少必需列：{missing}")
    frame = frame.loc[:, REQUIRED_COLUMNS].copy()
    if model_name.startswith("fm_heatmap_"):
        # 正式 heatmap 网格从 n=2 开始。这里额外过滤历史目录中的 n=1，避免
        # 旧别名或尚未重跑的结果把无定义规格重新带回最终 Excel 汇总。
        n1_parameter = frame["factor"].astype(str).str.contains(
            HEATMAP_N1_PARAMETER_PATTERN, regex=True, na=False
        )
        frame = frame.loc[~n1_parameter].copy()
    frame.insert(0, "model", model_name)
    frame["factor"] = frame["factor"].astype(str)
    frame["variable"] = frame["variable"].astype(str)
    frame["model_parameter"] = make_unique_parameter_labels(frame["factor"])
    frame["variable_main_name"] = [
        extract_variable_main_name(variable, factor, model_parameter)
        for variable, factor, model_parameter in zip(
            frame["variable"], frame["factor"], frame["model_parameter"]
        )
    ]

    # 少数交互模型会在同一个参数列中同时放入“普通 FAC”和“市态 FAC”。如果
    # 一律删除 hs300up/hs300down，两种不同解释变量会撞成同名行。仅对确实
    # 发生这种冲突的记录保留市态限定词，其余变量仍执行完整的参数清洗。
    collision_sizes = frame.groupby(
        ["model_parameter", "variable_main_name"], sort=False
    )["variable"].transform("nunique")
    collision_mask = collision_sizes.gt(1)
    if collision_mask.any():
        frame.loc[collision_mask, "variable_main_name"] = [
            extract_variable_main_name(
                variable,
                factor,
                model_parameter,
                remove_qualifier=False,
            )
            for variable, factor, model_parameter in frame.loc[
                collision_mask,
                ["variable", "factor", "model_parameter"],
            ].itertuples(index=False, name=None)
        ]
    frame["estimate"] = [
        format_estimate(coef, t_stat, p_value)
        for coef, t_stat, p_value in zip(
            frame["coef"], frame["t_stat"], frame["p_value"]
        )
    ]
    return frame


def extract_base_parameter(factor: str) -> str:
    """从冗长 factor 名中提取参数设置，例如 ``m3_n6_pairwise1``。"""
    match = PARAMETER_PATTERN.search(factor)
    if match:
        return match.group(0)
    horizon_match = HORIZON_PATTERN.search(factor)
    if horizon_match:
        return horizon_match.group(1)
    # 未知命名不能静默压成空值；保留原 factor 便于人工发现新口径。
    return factor


def remove_parameter_component(variable: str, component: str) -> str:
    """从变量名中删除一个完整参数段，同时保留其他下划线结构。"""
    if not component:
        return variable
    pattern = re.compile(rf"(?:^|_){re.escape(component)}(?=__|_|$)")
    return pattern.sub("", variable, count=1).lstrip("_")


def extract_variable_main_name(
    variable: str,
    factor: str,
    model_parameter: str,
    *,
    remove_qualifier: bool = True,
) -> str:
    """去掉变量名中已经由参数列表示的部分，得到用于合并行的主名称。

    例如，``FAC_rank_vol_m3_n6_pairwise1`` 会变成 ``FAC_rank_vol``；如果变量
    后面还有 ``__dmcs`` 这类真正属于变量定义的限定词，则只移除中间的参数，
    保留为 ``FAC_rank_vol__dmcs``。

    有些解释变量会把参数拆开放置，例如
    ``dummy_top50_hs300up_m3_n6_hit_above4_pairwise1``。此时分别删除参数列中
    已有的 ``hs300up``、``m3_n6`` 和 ``pairwise1``，得到
    ``dummy_top50_hit_above4``。未知参数格式不做猜测，避免误删变量名称。
    """
    factor_text = str(factor)
    variable_text = str(variable)
    parameter_label = str(model_parameter)

    # 只有脚本明确识别出的两类参数才参与清洗。extract_base_parameter 对未知
    # factor 会返回完整名称，不能用它直接替换，否则可能把整个变量名删空。
    parameter_match = PARAMETER_PATTERN.search(factor_text)
    if parameter_match:
        parameter = parameter_match.group(0)
        window_parameter, pairwise_parameter = parameter.rsplit("_", maxsplit=1)

        # 模型参数列可能还含 hs300up、hs300down、hitrate 等限定词。既然这些
        # 信息已经在列标题中出现，也应从解释变量主名称中删除。
        qualifier = ""
        if parameter_label.endswith(parameter):
            qualifier = parameter_label[: -len(parameter)].rstrip("_")

        main_name = variable_text
        components = [parameter, window_parameter, pairwise_parameter]
        if remove_qualifier:
            components.insert(0, qualifier)
        for component in components:
            main_name = remove_parameter_component(main_name, component)
        return main_name or variable_text
    else:
        horizon_match = HORIZON_PATTERN.search(factor_text)
        if horizon_match is None:
            return variable_text
        parameter = horizon_match.group(1)

    # 跨期限参数通常以完整后缀出现，按一个整体删除即可。
    main_name = remove_parameter_component(variable_text, parameter)
    return main_name or variable_text


def extract_parameter_qualifier(factor: str, base_parameter: str) -> str:
    """在参数发生重名时，提取区分市态或模型层所需的最短限定词。"""
    position = factor.find(base_parameter)
    prefix = factor[:position].rstrip("_") if position >= 0 else factor
    replacements = (
        ("controls_only_aligned_FAC_rank_vol", ""),
        ("FAC_rank_vol", ""),
        ("winrate_cumulative", "cumulative"),
        ("hitrate_top50", "hitrate"),
        ("hitrate_top33", "hitrate"),
        ("hitrate_bottom33", "hitrate"),
        ("winrate", ""),
        ("rank_vol", ""),
    )
    for source, replacement in replacements:
        if prefix == source:
            return replacement
        source_with_separator = f"{source}_"
        if prefix.startswith(source_with_separator):
            remainder = prefix[len(source_with_separator) :]
            return "_".join(part for part in (replacement, remainder) if part)
    return prefix


def make_unique_parameter_labels(factors: pd.Series) -> pd.Series:
    """生成模型内部唯一、但尽可能短的模型参数标签。

    普通模型直接得到 ``m3_n6_pairwise1``。只有多个 factor 会压成同一参数时，
    才追加必要限定词，例如 hs300up/hs300down 或 hitrate/cumulative。
    """
    unique_factors = list(dict.fromkeys(factors.astype(str)))
    base_by_factor = {factor: extract_base_parameter(factor) for factor in unique_factors}
    counts: dict[str, int] = {}
    for base in base_by_factor.values():
        counts[base] = counts.get(base, 0) + 1

    label_by_factor: dict[str, str] = {}
    for factor, base in base_by_factor.items():
        if counts[base] == 1:
            label_by_factor[factor] = base
            continue
        qualifier = extract_parameter_qualifier(factor, base)
        label_by_factor[factor] = f"{qualifier}_{base}" if qualifier else base

    duplicated_labels = {
        label for label in label_by_factor.values() if list(label_by_factor.values()).count(label) > 1
    }
    if duplicated_labels:
        details = {
            label: [factor for factor, value in label_by_factor.items() if value == label]
            for label in sorted(duplicated_labels)
        }
        raise ValueError(f"模型参数精简后仍不唯一：{details}")
    return factors.astype(str).map(label_by_factor)


def load_portfolio_summary(
    model_name: str,
    csv_path: Path,
) -> tuple[pd.DataFrame, str]:
    """读取并清洗一个模型最新的投资组合排序汇总。"""
    frame = pd.read_csv(csv_path)
    required_columns = {"sort_col", "portfolio", "consistency_mean"}
    missing = sorted(required_columns - set(frame.columns))
    if missing:
        raise ValueError(f"{csv_path} 缺少必需列：{missing}")

    future_return_columns = [
        column
        for column in frame.columns
        if re.fullmatch(r"future_ret_\d+m", str(column))
    ]
    if len(future_return_columns) != 1:
        raise ValueError(
            f"{csv_path} 应恰好包含一个 future_ret_*m 列，实际为："
            f"{future_return_columns}"
        )
    future_return_column = future_return_columns[0]

    frame = frame.loc[
        :, ["sort_col", "portfolio", "consistency_mean", future_return_column]
    ].copy()
    frame.insert(0, "model", model_name)
    frame["sort_col"] = frame["sort_col"].astype(str).str.strip()
    frame["portfolio"] = frame["portfolio"].astype(str).str.strip()
    frame["model_parameter"] = make_unique_parameter_labels(frame["sort_col"])
    frame["sort_variable_main"] = [
        extract_variable_main_name(sort_col, sort_col, parameter)
        for sort_col, parameter in zip(
            frame["sort_col"], frame["model_parameter"]
        )
    ]
    return frame, future_return_column


def build_portfolio_rows(
    frame: pd.DataFrame,
    future_return_column: str,
) -> tuple[list[str], list[list[object]]]:
    """删除独立 t/p 行，并把统计显著性折叠进 long-short 收益单元格。"""
    expected_portfolios = {
        "Q1",
        "Q2",
        "Q3",
        "Q4",
        "Q5",
        "long-short",
        "t-stat",
        "p-value",
    }
    rows: list[list[object]] = []
    for sort_col, group in frame.groupby("sort_col", sort=False):
        portfolio_labels = set(group["portfolio"])
        if portfolio_labels != expected_portfolios or group["portfolio"].duplicated().any():
            raise ValueError(
                f"{group['model'].iloc[0]} / {sort_col} 的组合行不完整或重复："
                f"{sorted(portfolio_labels)}"
            )

        by_portfolio = group.set_index("portfolio", drop=False)
        t_stat = by_portfolio.at["t-stat", future_return_column]
        p_value = by_portfolio.at["p-value", future_return_column]
        for portfolio in ("Q1", "Q2", "Q3", "Q4", "Q5", "long-short"):
            record = by_portfolio.loc[portfolio]
            future_return: object
            if portfolio == "long-short":
                future_return = format_portfolio_long_short(
                    record[future_return_column], t_stat, p_value
                )
            else:
                future_return = numeric_or_none(record[future_return_column])
            rows.append(
                [
                    record["model"],
                    record["sort_variable_main"],
                    record["model_parameter"],
                    portfolio,
                    numeric_or_none(record["consistency_mean"]),
                    future_return,
                ]
            )

    horizon = re.fullmatch(r"future_ret_(\d+)m", future_return_column)
    if horizon is None:
        raise ValueError(f"无法识别未来收益期限：{future_return_column}")
    headers = [
        "模型目录",
        "排序变量主名称",
        "模型参数",
        "投资组合",
        "一致性均值",
        f"未来{horizon.group(1)}个月收益率",
    ]
    return headers, rows


def build_comparison_rows(frame: pd.DataFrame) -> list[list[object]]:
    """每个模型目录和模型参数仅保留一行样本与拟合优度统计。"""
    statistic_columns = [
        "n_months",
        "avg_monthly_n",
        "avg_r_squared",
        "avg_adj_r_squared",
    ]
    rows: list[list[object]] = []
    for (model, parameter), group in frame.groupby(
        ["model", "model_parameter"], sort=False, dropna=False
    ):
        unique_statistics = group.loc[:, statistic_columns].drop_duplicates()
        if len(unique_statistics) != 1:
            raise ValueError(
                f"{model} / {parameter} 的样本统计不唯一，不能安全生成模型比较表"
            )
        statistics = unique_statistics.iloc[0]
        n_months = numeric_or_none(statistics["n_months"])
        rows.append(
            [
                model,
                parameter,
                int(n_months) if n_months is not None else None,
                numeric_or_none(statistics["avg_monthly_n"]),
                numeric_or_none(statistics["avg_r_squared"]),
                numeric_or_none(statistics["avg_adj_r_squared"]),
            ]
        )
    return rows


def build_significance_table(frame: pd.DataFrame) -> tuple[list[str], list[list[object]]]:
    """按变量主名称合并行，并把模型参数作为横向列。"""
    parameters = list(dict.fromkeys(frame["model_parameter"].astype(str)))
    row_keys = list(
        dict.fromkeys(
            zip(
                frame["model"].astype(str),
                frame["variable_main_name"].astype(str),
            )
        )
    )
    value_map: dict[tuple[str, str, str], str] = {}
    value_columns = [
        "model",
        "model_parameter",
        "variable_main_name",
        "variable",
        "estimate",
    ]
    for record in frame[value_columns].to_dict(orient="records"):
        key = (
            record["model"],
            record["model_parameter"],
            record["variable_main_name"],
        )
        if key in value_map and value_map[key] != record["estimate"]:
            raise ValueError(
                "变量去掉参数后缀后出现冲突系数："
                f"{key}；请检查原变量 {record['variable']}"
            )
        value_map[key] = record["estimate"]

    rows = [
        [model, variable_main_name]
        + [
            value_map.get((model, parameter, variable_main_name))
            for parameter in parameters
        ]
        for model, variable_main_name in row_keys
    ]
    return ["模型目录", "变量", *parameters], rows


def group_result_files(
    result_files: dict[str, Path],
    model_groups: OrderedDict[str, list[str]],
) -> OrderedDict[str, list[tuple[str, Path]]]:
    """按模型地图分配结果文件，未映射文件统一放入最后一个 Sheet。"""
    grouped: OrderedDict[str, list[tuple[str, Path]]] = OrderedDict()
    mapped_models: set[str] = set()

    for sheet_title, model_order in model_groups.items():
        items = [
            (model_name, result_files[model_name])
            for model_name in model_order
            if model_name in result_files
        ]
        grouped[sheet_title] = items
        mapped_models.update(model_name for model_name, _ in items)

    unmapped = [
        (model_name, result_files[model_name])
        for model_name in sorted(set(result_files) - mapped_models)
    ]
    if unmapped:
        grouped[UNMAPPED_SHEET_TITLE] = unmapped
    return grouped


def make_output_sheet_name(base_title: str, suffix: str) -> str:
    """生成带指定后缀且不超过 Excel 31 字符限制的 Sheet 名。"""
    available_length = 31 - len(suffix)
    if available_length <= 0:
        raise ValueError(f"Sheet 后缀过长：{suffix}")
    return clean_family_title(base_title)[:available_length] + suffix


def apply_academic_table_style(
    worksheet,
    *,
    title: str,
    headers: list[str] | tuple[str, ...],
    first_data_row: int,
    last_data_row: int,
    table_kind: str,
) -> None:
    """应用学术三线表样式：顶线、表头底线和表格底线，无竖线。"""
    last_column = len(headers)
    last_column_letter = get_column_letter(last_column)
    header_row = first_data_row - 1
    # 参数很多的显著性表可能超过 100 列；若标题横跨全部列，居中位置会跑到
    # 屏幕很远处，打开左上角时反而看不到标题。标题区最多使用前 10 列。
    title_last_column = min(last_column, 10)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = f"C{first_data_row}" if table_kind == "significance" else f"A{first_data_row}"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_last_column)
    worksheet.cell(1, 1, title)
    set_bilingual_text(worksheet.cell(1, 1), title, size=14, bold=True)
    worksheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 25

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=title_last_column)
    note = (
        "注：括号内为 Newey-West t 值；***、**、* 分别表示在 1%、5%、10% 水平显著。"
        if table_kind == "significance"
        else "注：每行对应一个模型目录和唯一模型参数；R² 与调整 R² 以百分比显示。"
    )
    worksheet.cell(2, 1, note)
    set_bilingual_text(
        worksheet.cell(2, 1), note, size=9, italic=True, color="666666"
    )
    worksheet.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[2].height = 19

    medium_black = Side(style="medium", color="000000")
    thin_black = Side(style="thin", color="000000")
    header_border = Border(top=medium_black, bottom=thin_black)
    bottom_border = Border(bottom=medium_black)

    for cell in worksheet[header_row]:
        set_bilingual_text(cell, str(cell.value), size=10, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    worksheet.row_dimensions[header_row].height = 31

    for row in worksheet.iter_rows(
        min_row=first_data_row,
        max_row=last_data_row,
        min_col=1,
        max_col=last_column,
    ):
        for column_index, cell in enumerate(row, start=1):
            set_bilingual_text(cell, cell.value, size=9)
            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if table_kind == "significance" and column_index >= 3
                    else "right"
                    if table_kind == "comparison" and column_index >= 3
                    else "left"
                ),
                vertical="center",
                wrap_text=True,
            )
        worksheet.row_dimensions[row[0].row].height = 31 if table_kind == "significance" else 22

    if table_kind == "comparison":
        # 统计量保留两位小数；月份数是计数，不显示无意义的小数位。
        for row_index in range(first_data_row, last_data_row + 1):
            worksheet.cell(row_index, 3).number_format = "#,##0"
            worksheet.cell(row_index, 4).number_format = "#,##0.00"
            worksheet.cell(row_index, 5).number_format = "0.00%"
            worksheet.cell(row_index, 6).number_format = "0.00%"

    for cell in worksheet[last_data_row]:
        cell.border = bottom_border

    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 32
    if table_kind == "comparison":
        for column_letter, width in {"C": 11, "D": 15, "E": 13, "F": 16}.items():
            worksheet.column_dimensions[column_letter].width = width
    else:
        for column_index, header in enumerate(headers[2:], start=3):
            # 回归变量名可能很长，适当放宽但设置上限，避免单列无限膨胀。
            width = max(15, min(38, len(str(header)) * 1.05 + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = f"1:{header_row}"
    worksheet.print_area = f"A1:{last_column_letter}{last_data_row}"
    worksheet.sheet_properties.outlinePr.summaryBelow = True


def make_unique_model_sheet_name(model_name: str, used_names: set[str]) -> str:
    """把模型目录名转换为不超过 31 字符且不重名的 Sheet 名。"""
    # clean_family_title 会立即截断到 31 字符，不适合这里“保留名称尾部”的需求，
    # 因此先只做非法字符与空白清洗，最后再由 compact_name 控制长度。
    base_name = html.unescape(model_name).strip()
    base_name = re.sub(r"[\\/*?:\[\]]", "_", base_name)
    base_name = re.sub(r"\s+", "", base_name) or "未命名模型"

    def compact_name(max_length: int) -> str:
        """长名称同时保留开头和结尾，便于识别 y1m、hs300 等关键尾缀。"""
        if len(base_name) <= max_length:
            return base_name
        tail_length = min(12, max_length // 2)
        head_length = max_length - tail_length - 1
        return f"{base_name[:head_length]}~{base_name[-tail_length:]}"

    sheet_name = compact_name(31)
    suffix_number = 2
    while sheet_name in used_names:
        suffix = f"_{suffix_number}"
        sheet_name = f"{compact_name(31 - len(suffix))}{suffix}"
        suffix_number += 1
    used_names.add(sheet_name)
    return sheet_name


def apply_portfolio_table_style(
    worksheet,
    *,
    model_name: str,
    source_file_name: str,
    headers: list[str],
    first_data_row: int,
    last_data_row: int,
) -> None:
    """为单模型投资组合排序结果应用与回归汇总一致的三线表样式。"""
    last_column = len(headers)
    last_column_letter = get_column_letter(last_column)
    header_row = first_data_row - 1

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = f"D{first_data_row}"
    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=last_column
    )
    title = f"Portfolio Sorting｜{model_name}"
    worksheet.cell(1, 1, title)
    set_bilingual_text(worksheet.cell(1, 1), title, size=14, bold=True)
    worksheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 25

    worksheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=last_column
    )
    note = (
        f"来源：{source_file_name}（该模型编号最新的 summary）；收益率按百分比显示。"
        "long-short 括号内为 t 值；***、**、* 分别表示在 1%、5%、10% 水平显著。"
    )
    worksheet.cell(2, 1, note)
    set_bilingual_text(
        worksheet.cell(2, 1), note, size=9, italic=True, color="666666"
    )
    worksheet.cell(2, 1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    worksheet.row_dimensions[2].height = 30

    medium_black = Side(style="medium", color="000000")
    thin_black = Side(style="thin", color="000000")
    light_gray = Side(style="thin", color="D9D9D9")
    header_border = Border(top=medium_black, bottom=thin_black)
    bottom_border = Border(bottom=medium_black)

    for cell in worksheet[header_row]:
        set_bilingual_text(cell, str(cell.value), size=10, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    worksheet.row_dimensions[header_row].height = 31

    previous_parameter: tuple[object, object] | None = None
    for row_index in range(first_data_row, last_data_row + 1):
        current_parameter = (
            worksheet.cell(row_index, 2).value,
            worksheet.cell(row_index, 3).value,
        )
        starts_new_block = current_parameter != previous_parameter
        for column_index in range(1, last_column + 1):
            cell = worksheet.cell(row_index, column_index)
            set_bilingual_text(cell, cell.value, size=9)
            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if column_index == 4 or (column_index == 6 and isinstance(cell.value, str))
                    else "right"
                    if column_index >= 5
                    else "left"
                ),
                vertical="center",
                wrap_text=True,
            )
            if starts_new_block and row_index > first_data_row:
                cell.border = Border(top=light_gray)
        worksheet.row_dimensions[row_index].height = (
            31 if worksheet.cell(row_index, 4).value == "long-short" else 22
        )
        previous_parameter = current_parameter

    # Q1-Q5 和多空组合的一致性均值均为比例；未来收益率的普通组合也是比例。
    for row_index in range(first_data_row, last_data_row + 1):
        worksheet.cell(row_index, 5).number_format = "0.00%"
        if worksheet.cell(row_index, 4).value != "long-short":
            worksheet.cell(row_index, 6).number_format = "0.00%"

    for cell in worksheet[last_data_row]:
        cell.border = bottom_border

    for column_letter, width in {
        "A": 42,
        "B": 34,
        "C": 28,
        "D": 14,
        "E": 16,
        "F": 20,
    }.items():
        worksheet.column_dimensions[column_letter].width = width

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = f"1:{header_row}"
    worksheet.print_area = f"A1:{last_column_letter}{last_data_row}"


def write_portfolio_workbook(
    latest_files: dict[str, Path],
    model_groups: OrderedDict[str, list[str]],
    output_path: Path,
) -> tuple[int, int, int]:
    """按模型地图顺序为每个有结果的模型生成一张投资组合排序 Sheet。"""
    model_order = [
        model_name
        for model_names in model_groups.values()
        for model_name in model_names
        if model_name in latest_files
    ]
    mapped_models = set(model_order)
    model_order.extend(sorted(set(latest_files) - mapped_models))

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.loaded_theme = build_bilingual_theme()
    used_sheet_names: set[str] = set()
    total_rows = 0

    for model_name in model_order:
        csv_path = latest_files[model_name]
        frame, future_return_column = load_portfolio_summary(model_name, csv_path)
        headers, rows = build_portfolio_rows(frame, future_return_column)
        sheet_name = make_unique_model_sheet_name(model_name, used_sheet_names)
        worksheet = workbook.create_sheet(title=sheet_name)

        header_row = 4
        first_data_row = header_row + 1
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(header_row, column_index, header)
        for row in rows:
            worksheet.append(row)

        last_data_row = worksheet.max_row
        apply_portfolio_table_style(
            worksheet,
            model_name=model_name,
            source_file_name=csv_path.name,
            headers=headers,
            first_data_row=first_data_row,
            last_data_row=last_data_row,
        )
        total_rows += len(rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(model_order), total_rows, len(workbook.sheetnames)


def write_workbook(
    grouped_files: OrderedDict[str, list[tuple[str, Path]]],
    output_path: Path,
) -> tuple[int, int, int, int]:
    """每个模型地图分组写出两类 Sheet，并返回处理摘要。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.loaded_theme = build_bilingual_theme()

    total_models = 0
    total_comparison_rows = 0
    total_coefficient_cells = 0
    for base_title, model_files in grouped_files.items():
        frames: list[pd.DataFrame] = []
        for model_name, csv_path in model_files:
            frames.append(load_result_frame(model_name, csv_path))
            total_models += 1

        if frames:
            combined = pd.concat(frames, ignore_index=True)
            comparison_rows = build_comparison_rows(combined)
            significance_headers, significance_rows = build_significance_table(combined)
        else:
            comparison_rows = []
            significance_headers = ["模型目录", "模型参数"]
            significance_rows = []

        sheet_specs = (
            (
                make_output_sheet_name(base_title, "_模型比较"),
                list(COMPARISON_HEADERS),
                comparison_rows,
                "comparison",
            ),
            (
                make_output_sheet_name(base_title, "_回归系数显著性"),
                significance_headers,
                significance_rows,
                "significance",
            ),
        )

        for sheet_name, headers, rows, table_kind in sheet_specs:
            worksheet = workbook.create_sheet(title=sheet_name)
            header_row = 4
            first_data_row = header_row + 1
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(header_row, column_index, header)

            if rows:
                for row in rows:
                    worksheet.append(row)
            else:
                worksheet.append(["本分组暂无结果文件", *([None] * (len(headers) - 1))])

            last_data_row = worksheet.max_row
            apply_academic_table_style(
                worksheet,
                title=f"Fama-MacBeth｜{sheet_name}",
                headers=headers,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                table_kind=table_kind,
            )

        total_comparison_rows += len(comparison_rows)
        total_coefficient_cells += int(combined.shape[0]) if frames else 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return total_models, total_comparison_rows, total_coefficient_cells, len(workbook.sheetnames)


def main() -> None:
    """生成回归汇总和投资组合排序汇总，并打印便于核对的摘要。"""
    args = parse_args()
    input_root = args.input_root.resolve()
    model_map_path = args.model_map.resolve()
    output_path = args.output.resolve()
    portfolio_output_path = args.portfolio_output.resolve()

    model_groups = parse_model_groups(model_map_path)
    result_files = discover_result_files(input_root)
    grouped_files = group_result_files(result_files, model_groups)
    model_count, comparison_row_count, coefficient_cell_count, sheet_count = write_workbook(
        grouped_files, output_path
    )

    mapped_models = {
        model_name
        for sheet_title, items in grouped_files.items()
        if sheet_title != UNMAPPED_SHEET_TITLE
        for model_name, _ in items
    }
    unmapped_models = sorted(set(result_files) - mapped_models)

    print(f"已生成：{output_path}")
    print(
        f"共处理 {model_count} 个模型、{comparison_row_count} 行模型比较、"
        f"{coefficient_cell_count} 个系数单元、{sheet_count} 个 Sheet。"
    )
    if unmapped_models:
        print("以下模型未出现在 model_map.html，已放入‘未映射模型’Sheet：")
        for model_name in unmapped_models:
            print(f"- {model_name}")

    latest_portfolio_files = discover_latest_portfolio_summaries(input_root)
    portfolio_model_count, portfolio_row_count, portfolio_sheet_count = (
        write_portfolio_workbook(
            latest_portfolio_files,
            model_groups,
            portfolio_output_path,
        )
    )
    print(f"已生成：{portfolio_output_path}")
    print(
        f"共处理 {portfolio_model_count} 个模型的最新投资组合排序结果、"
        f"{portfolio_row_count} 行清洗后数据、{portfolio_sheet_count} 个 Sheet。"
    )


if __name__ == "__main__":
    main()
