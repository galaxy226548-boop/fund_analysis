"""验证 Fama-MacBeth 学术汇总表的分组、格式化与完整性。"""

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl.cell.rich_text import CellRichText
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = PROJECT_ROOT / "D_analysis" / "scripts" / "result_sorting.py"


def load_script_module():
    """按文件路径加载脚本，避免 scripts 目录必须声明成 Python 包。"""
    spec = importlib.util.spec_from_file_location("result_sorting_test", SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise ImportError(f"无法加载脚本：{SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


result_sorting = load_script_module()


class ResultSortingTests(unittest.TestCase):
    """覆盖星号、模型地图解析和工作簿三线表输出。"""

    def test_estimate_format_uses_two_decimals_t_stat_and_stars(self) -> None:
        self.assertEqual(result_sorting.format_estimate(-0.004, 2.345, 0.009), "0.00***\n(2.35)")
        self.assertEqual(result_sorting.format_estimate(1.236, -1.8, 0.049), "1.24**\n(-1.80)")
        self.assertEqual(result_sorting.format_estimate(0.1, 1.65, 0.099), "0.10*\n(1.65)")
        self.assertEqual(result_sorting.format_estimate(0.1, 0.2, 0.5), "0.10\n(0.20)")

    def test_real_model_map_has_all_groups_and_registry_models(self) -> None:
        groups = result_sorting.parse_model_groups(result_sorting.DEFAULT_MODEL_MAP)
        model_keys = [key for keys in groups.values() for key in keys]
        self.assertEqual(len(groups), 7)
        self.assertEqual(len(model_keys), 103)
        self.assertEqual(len(model_keys), len(set(model_keys)))
        self.assertIn("fm_null", model_keys)
        self.assertIn("fm_ymatch_cross_indvol_lowvol", model_keys)

    def test_parameter_labels_are_short_and_preserve_required_qualifiers(self) -> None:
        ordinary = pd.Series(["FAC_rank_vol_m3_n6_pairwise1"])
        self.assertEqual(
            result_sorting.make_unique_parameter_labels(ordinary).tolist(),
            ["m3_n6_pairwise1"],
        )

        states = pd.Series(
            [
                "FAC_rank_vol_hs300up_m3_n6_pairwise1",
                "FAC_rank_vol_hs300down_m3_n6_pairwise1",
            ]
        )
        self.assertEqual(
            result_sorting.make_unique_parameter_labels(states).tolist(),
            ["hs300up_m3_n6_pairwise1", "hs300down_m3_n6_pairwise1"],
        )

        layers = pd.Series(
            [
                "hitrate_top50_m1_n2_pairwise1",
                "winrate_cumulative_m1_n2_pairwise1",
            ]
        )
        self.assertEqual(
            result_sorting.make_unique_parameter_labels(layers).tolist(),
            ["hitrate_m1_n2_pairwise1", "cumulative_m1_n2_pairwise1"],
        )

        self.assertEqual(
            result_sorting.extract_base_parameter(
                "rank_vol_across_horizons_1m_3m_6m_12m"
            ),
            "1m_3m_6m_12m",
        )

    def test_heatmap_results_drop_legacy_n1_parameters(self) -> None:
        """历史 heatmap CSV 即使仍含 n=1，也不能进入最终汇总表。"""
        with tempfile.TemporaryDirectory() as temporary_directory:
            csv_path = Path(temporary_directory) / "fama_macbeth_results.csv"
            rows = []
            for n in (1, 2):
                factor = f"FAC_rank_vol_m3_n{n}_pairwise1"
                rows.append(
                    {
                        "factor": factor,
                        "variable": factor,
                        "coef": 0.1,
                        "t_stat": 2.0,
                        "p_value": 0.05,
                        "n_months": 100,
                        "avg_monthly_n": 200,
                        "avg_r_squared": 0.2,
                        "avg_adj_r_squared": 0.1,
                    }
                )
            pd.DataFrame(rows).to_csv(csv_path, index=False)

            frame = result_sorting.load_result_frame("fm_heatmap_bm33", csv_path)
            self.assertEqual(frame["model_parameter"].tolist(), ["m3_n2_pairwise1"])
            self.assertFalse(frame["factor"].str.contains("_n1_").any())

    def test_workbook_has_academic_lines_percentages_and_unmapped_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            input_root = root / "inputs"
            model_dir = input_root / "fm_null"
            legacy_dir = input_root / "fm_legacy"
            model_dir.mkdir(parents=True)
            legacy_dir.mkdir(parents=True)

            rows = [
                {
                    "factor": "FAC_rank_vol_m3_n6_pairwise1",
                    "variable": variable,
                    "coef": coef,
                    "newey_west_se": 0.01,
                    "t_stat": t_stat,
                    "p_value": p_value,
                    "n_months": 79,
                    "avg_monthly_n": 197.164557,
                    "total_regression_obs": 15576,
                    "avg_r_squared": 0.278487,
                    "avg_adj_r_squared": 0.235387,
                    "newey_west_lag": 5,
                    "min_cross_section_n": 50,
                }
                for variable, coef, t_stat, p_value in (
                    ("CtrlRetSTR", 0.1234, 2.345, 0.009),
                    ("as_偏股混合型基金", -0.2, -1.8, 0.049),
                )
            ]
            pd.DataFrame(rows).to_csv(model_dir / result_sorting.RESULT_FILE_NAME, index=False)
            pd.DataFrame(rows).to_csv(legacy_dir / result_sorting.RESULT_FILE_NAME, index=False)

            groups = result_sorting.OrderedDict([("1_基础模型", ["fm_null"])])
            files = result_sorting.discover_result_files(input_root)
            grouped = result_sorting.group_result_files(files, groups)
            output = root / "summary.xlsx"
            model_count, comparison_rows, coefficient_cells, sheet_count = (
                result_sorting.write_workbook(grouped, output)
            )

            self.assertEqual(model_count, 2)
            self.assertEqual(comparison_rows, 2)
            self.assertEqual(coefficient_cells, 4)
            self.assertEqual(sheet_count, 4)
            workbook = load_workbook(output, rich_text=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "1_基础模型_模型比较",
                    "1_基础模型_回归系数显著性",
                    "未映射模型_模型比较",
                    "未映射模型_回归系数显著性",
                ],
            )

            comparison = workbook["1_基础模型_模型比较"]
            self.assertEqual(comparison["A5"].value, "fm_null")
            self.assertEqual(comparison["B5"].value, "m3_n6_pairwise1")
            self.assertEqual(comparison["C5"].value, 79)
            self.assertAlmostEqual(comparison["D5"].value, 197.164557)
            self.assertEqual(comparison["E5"].number_format, "0.00%")
            self.assertEqual(comparison["F5"].number_format, "0.00%")
            self.assertEqual(comparison["A4"].border.top.style, "medium")
            self.assertEqual(comparison["A4"].border.bottom.style, "thin")
            self.assertEqual(comparison["A5"].border.bottom.style, "medium")

            significance = workbook["1_基础模型_回归系数显著性"]
            self.assertEqual(significance["A5"].value, "fm_null")
            self.assertEqual(significance["B5"].value, "CtrlRetSTR")
            self.assertEqual(significance["C4"].value, "m3_n6_pairwise1")
            self.assertEqual(significance["C5"].value, "0.12***\n(2.35)")
            self.assertEqual(str(significance["B6"].value), "as_偏股混合型基金")
            self.assertEqual(significance["C6"].value, "-0.20**\n(-1.80)")

            # 纯英文、纯中文和混排文本都必须写入明确的字体记录。
            self.assertEqual(comparison["A5"].font.name, "Times New Roman")
            self.assertEqual(comparison["A4"].font.name, "微软雅黑")
            self.assertIsInstance(significance["B6"].value, CellRichText)
            mixed_fonts = {
                block.font.rFont
                for block in significance["B6"].value
                if hasattr(block, "font")
            }
            self.assertEqual(mixed_fonts, {"Times New Roman", "微软雅黑"})

            # 主题字体继续作为不支持富文本环境的回退。
            with ZipFile(output) as archive:
                theme = archive.read("xl/theme/theme1.xml").decode("utf-8")
            self.assertIn('<a:latin typeface="Times New Roman"/>', theme)
            self.assertIn('<a:ea typeface="微软雅黑"/>', theme)
            self.assertIn('<a:font script="Hans" typeface="微软雅黑"/>', theme)


if __name__ == "__main__":
    unittest.main()
