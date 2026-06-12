import sys
import streamlit as st
import pandas as pd
from pathlib import Path

# 让 Python 能找到 src/mapping_loader.py（与本文件同级的 src/ 目录）
sys.path.insert(0, str(Path(__file__).parent))
from src.mapping_loader import load_mapping, build_file_coverage

# ── 默认路径常量 ──────────────────────────────────────────────────────────────
DEFAULT_RAW_DIR   = r"/Users/chloezh/Projects/Fund_Analysis/A_data/data"
DEFAULT_CLEAN_DIR   = r"/Users/chloezh/Projects/Fund_Analysis/A_data/prepared_data"
DEFAULT_MAPPING_JSON = r"/Users/chloezh/Projects/Fund_Analysis/A_data/reference/data_inventory_A.json"

st.set_page_config(page_title="清洗覆盖审计", page_icon="🗂️", layout="wide")
st.title("🗂️ Parquet 清洗覆盖审计")

# ── session_state 初始化 ──────────────────────────────────────────────────────
if "scanned" not in st.session_state:
    st.session_state.scanned = False
if "raw_dir_saved" not in st.session_state:
    st.session_state.raw_dir_saved = ""
if "cleaned_dir_saved" not in st.session_state:
    st.session_state.cleaned_dir_saved = ""
if "mapping_json_saved" not in st.session_state:
    st.session_state.mapping_json_saved = ""

# ── 侧边栏 ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("目录配置")
    raw_dir_input = st.text_input(
        "raw_dir（原始数据目录）", placeholder="/data/raw",
        value=st.session_state.raw_dir_saved or DEFAULT_RAW_DIR,
    )
    cleaned_dir_input = st.text_input(
        "cleaned_dir（清洗数据目录）", placeholder="/data/cleaned",
        value=st.session_state.cleaned_dir_saved or DEFAULT_CLEAN_DIR,
    )
    mapping_json_input = st.text_input(
        "mapping JSON（可选）", placeholder="/path/to/mapping.json",
        value=st.session_state.mapping_json_saved or DEFAULT_MAPPING_JSON,
    )
    if st.button("🔍 扫描", use_container_width=True):
        st.session_state.scanned            = True
        st.session_state.raw_dir_saved      = raw_dir_input
        st.session_state.cleaned_dir_saved  = cleaned_dir_input
        st.session_state.mapping_json_saved = mapping_json_input

# ── 未扫描时提示 ──────────────────────────────────────────────────────────────
if not st.session_state.scanned:
    st.info("请在左侧填写两个目录路径，然后点击「扫描」。")
    st.stop()

# ── 使用 session_state 中保存的路径 ──────────────────────────────────────────
raw_str     = st.session_state.raw_dir_saved.strip()
cleaned_str = st.session_state.cleaned_dir_saved.strip()

raw_dir     = Path(raw_str).expanduser()     if raw_str     else None
cleaned_dir = Path(cleaned_str).expanduser() if cleaned_str else None

errors = []
if not raw_dir or not raw_dir.exists():
    errors.append(f"❌ raw_dir 不存在或未填写：`{raw_str}`")
if not cleaned_dir or not cleaned_dir.exists():
    errors.append(f"❌ cleaned_dir 不存在或未填写：`{cleaned_str}`")

if errors:
    for e in errors:
        st.error(e)
    st.stop()

# ── 扫描文件 ──────────────────────────────────────────────────────────────────
SCAN_SUFFIXES = {".parquet", ".xlsx", ".csv"}
raw_map     = {p.name: p for p in raw_dir.rglob("*")
               if p.suffix in SCAN_SUFFIXES and "old_data_archive" not in p.parts}
cleaned_map = {p.name: p for p in cleaned_dir.rglob("*") if p.suffix in SCAN_SUFFIXES}
all_files   = raw_map.keys() | cleaned_map.keys()

# ── 加载 mapping（可选）──────────────────────────────────────────────────────
mapping      = None
mapping_note = None
mapping_str  = st.session_state.mapping_json_saved.strip()
if mapping_str:
    mapping, mapping_err = load_mapping(mapping_str)
    if mapping_err:
        st.warning(f"⚠️ mapping JSON 读取失败，已回退到同名匹配：{mapping_err}")
    else:
        mapping_note = f"✅ 已加载 mapping JSON，共 {len(mapping)} 条记录"

if mapping_note:
    st.info(mapping_note)

# ── 覆盖表 ────────────────────────────────────────────────────────────────────
rows = build_file_coverage(raw_map, cleaned_map, mapping)
df   = pd.DataFrame(rows, columns=[
    "file_name", "raw_exists", "cleaned_exists", "status",
    "match_method", "mapped_raw_files", "mapping_status",
])

# ── 统计指标 ──────────────────────────────────────────────────────────────────
n_raw       = len(raw_map)
n_cleaned   = len(cleaned_map)
n_uncleaned = int((df["status"] == "⏳ 未清洗").sum())
n_orphan    = int((df["status"] == "👻 孤儿清洗文件").sum())

c1, c2, c3, c4 = st.columns(4)
c1.metric("Raw 文件数",    n_raw)
c2.metric("Cleaned 文件数", n_cleaned)
c3.metric("未清洗",        n_uncleaned, delta=f"-{n_uncleaned}" if n_uncleaned else None, delta_color="inverse")
c4.metric("孤儿清洗文件",  n_orphan,    delta=f"{n_orphan}"     if n_orphan    else None, delta_color="inverse")

st.divider()

# ── 过滤器 + 覆盖表 ───────────────────────────────────────────────────────────
status_options  = ["全部"] + sorted(df["status"].unique().tolist())
selected_status = st.selectbox("按状态筛选", status_options)
display_df      = df if selected_status == "全部" else df[df["status"] == selected_status]

st.subheader(f"覆盖表（共 {len(display_df)} 条）")
st.dataframe(
    display_df, use_container_width=True, hide_index=True,
    column_config={
        "file_name":        st.column_config.TextColumn("文件名（cleaned）", width="large"),
        "raw_exists":       st.column_config.CheckboxColumn("raw 存在"),
        "cleaned_exists":   st.column_config.CheckboxColumn("cleaned 存在"),
        "status":           st.column_config.TextColumn("状态",          width="medium"),
        "match_method":     st.column_config.TextColumn("匹配方式",       width="small"),
        "mapped_raw_files": st.column_config.TextColumn("对应 raw 文件",  width="large"),
        "mapping_status":   st.column_config.TextColumn("mapping 状态",   width="medium"),
    },
)

st.divider()

# ── 辅助函数 ──────────────────────────────────────────────────────────────────

def build_profile(data: pd.DataFrame) -> pd.DataFrame:
    records = []
    for col in data.columns:
        s      = data[col]
        is_num = pd.api.types.is_numeric_dtype(s)
        records.append({
            "column":  col,
            "dtype":   str(s.dtype),
            "null_%":  round(s.isna().mean() * 100, 2),
            "min":     round(float(s.min()),  4) if is_num and s.notna().any() else None,
            "max":     round(float(s.max()),  4) if is_num and s.notna().any() else None,
            "mean":    round(float(s.mean()), 4) if is_num and s.notna().any() else None,
        })
    return pd.DataFrame(records)

def show_profile(data: pd.DataFrame, label: str):
    st.markdown(f"**{label}** — {data.shape[0]:,} 行 × {data.shape[1]} 列")
    prof = build_profile(data)
    st.dataframe(
        prof, use_container_width=True, hide_index=True,
        column_config={
            "column": st.column_config.TextColumn("字段名"),
            "dtype":  st.column_config.TextColumn("类型"),
            "null_%": st.column_config.ProgressColumn("缺失率 %", min_value=0, max_value=100, format="%.2f%%"),
            "min":    st.column_config.NumberColumn("min",  format="%.4f"),
            "max":    st.column_config.NumberColumn("max",  format="%.4f"),
            "mean":   st.column_config.NumberColumn("mean", format="%.4f"),
        },
    )

def show_profile_diff(raw_data: pd.DataFrame, cleaned_data: pd.DataFrame):
    raw_prof     = build_profile(raw_data).set_index("column")
    cleaned_prof = build_profile(cleaned_data).set_index("column")
    all_cols     = raw_prof.index.union(cleaned_prof.index)
    rows = []
    for col in all_cols:
        r = raw_prof.loc[col]     if col in raw_prof.index     else {}
        c = cleaned_prof.loc[col] if col in cleaned_prof.index else {}
        rows.append({
            "字段名":     col,
            "raw dtype":  r.get("dtype",  "—"),
            "cln dtype":  c.get("dtype",  "—"),
            "raw null %": r.get("null_%", None),
            "cln null %": c.get("null_%", None),
            "raw min":    r.get("min",    None),
            "cln min":    c.get("min",    None),
            "raw max":    r.get("max",    None),
            "cln max":    c.get("max",    None),
            "raw mean":   r.get("mean",   None),
            "cln mean":   c.get("mean",   None),
        })
    diff_df = pd.DataFrame(rows)
    st.dataframe(
        diff_df, use_container_width=True, hide_index=True,
        column_config={
            "字段名":      st.column_config.TextColumn("字段名",    width="medium"),
            "raw dtype":  st.column_config.TextColumn("raw 类型"),
            "cln dtype":  st.column_config.TextColumn("cln 类型"),
            "raw null %": st.column_config.ProgressColumn("raw 缺失%", min_value=0, max_value=100, format="%.2f%%"),
            "cln null %": st.column_config.ProgressColumn("cln 缺失%", min_value=0, max_value=100, format="%.2f%%"),
            "raw min":    st.column_config.NumberColumn("raw min",  format="%.4f"),
            "cln min":    st.column_config.NumberColumn("cln min",  format="%.4f"),
            "raw max":    st.column_config.NumberColumn("raw max",  format="%.4f"),
            "cln max":    st.column_config.NumberColumn("cln max",  format="%.4f"),
            "raw mean":   st.column_config.NumberColumn("raw mean", format="%.4f"),
            "cln mean":   st.column_config.NumberColumn("cln mean", format="%.4f"),
        },
    )

# ── 文件选择 & 预览 ───────────────────────────────────────────────────────────
st.subheader("📄 文件预览 & Profile")

# ── 步骤1：先选侧 ─────────────────────────────────────────────────────────────
search_side = st.radio(
    "① 选择要浏览的数据侧",
    ["Clean Data", "Raw Data"],
    horizontal=True,
    help="选定后，下方文件列表只显示该侧存在的文件",
)
browse_raw = (search_side == "Raw Data")

# ── 步骤2：根据所选侧生成文件列表 ────────────────────────────────────────────
if browse_raw:
    # raw 侧：用 mapped_raw_files 里的文件名（去重、排序）
    raw_candidates = set()
    for _, row in df[df["raw_exists"] == True].iterrows():
        mapped = row.get("mapped_raw_files", "")
        if mapped and mapped != "—":
            for fn in mapped.split(","):
                fn = fn.strip()
                if fn and fn in raw_map:
                    raw_candidates.add(fn)
        else:
            if row["file_name"] in raw_map:
                raw_candidates.add(row["file_name"])
    file_options = ["（请选择）"] + sorted(raw_candidates)
else:
    # clean 侧：直接用 df 的 file_name（cleaned_exists=True 的）
    clean_candidates = sorted(df[df["cleaned_exists"] == True]["file_name"].tolist())
    file_options = ["（请选择）"] + clean_candidates

selected_file = st.selectbox("② 选择要检视的文件", file_options)

if not selected_file or selected_file == "（请选择）":
    st.stop()

# ── 根据所选侧定位 file_row、raw_path、cleaned_path ──────────────────────────
if browse_raw:
    # 以 raw 文件名反查 df 里对应的行（mapped_raw_files 包含该文件名）
    matched_rows = df[df["mapped_raw_files"].str.contains(
        selected_file.replace(".", r"\."), regex=True, na=False
    )]
    if matched_rows.empty:
        # 同名兜底
        matched_rows = df[df["file_name"] == selected_file]
    file_row  = matched_rows.iloc[0] if not matched_rows.empty else None
    has_raw   = True
    has_clean = (file_row is not None and bool(file_row["cleaned_exists"]))
else:
    file_row  = df[df["file_name"] == selected_file].iloc[0]
    has_raw   = bool(file_row["raw_exists"])
    has_clean = bool(file_row["cleaned_exists"])

# ── 定位 raw 文件真实路径（mapping 模式下 raw 文件名可能与 file_name 不同）──
def find_raw_path(file_row, raw_map) -> Path | None:
    """优先用 mapped_raw_files 的第一个文件名，回退到 file_name 同名查找。"""
    mapped = file_row.get("mapped_raw_files", "")
    if mapped and mapped != "—":
        first = mapped.split(",")[0].strip()
        if first in raw_map:
            return raw_map[first]
    # 同名兜底
    name = file_row["file_name"]
    return raw_map.get(name)

def find_cleaned_path(file_row, cleaned_map) -> Path | None:
    name = file_row["file_name"]
    return cleaned_map.get(name)

# ── 根据后缀选择读取方式 ──────────────────────────────────────────────────────
PREVIEW_N  = 50    # 前/后各显示行数
SAMPLE_N   = 5000  # profile 用样本行数

def _xlsx_row_count(path: Path) -> int:
    """用 openpyxl read-only 模式快速获取行数（不解析单元格值）。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    n  = ws.max_row or 0
    wb.close()
    return max(0, n - 1)  # 减去表头行

@st.cache_data(show_spinner=False)
def read_head(path_str: str, n: int = PREVIEW_N):
    path = Path(path_str)
    try:
        if path.suffix == ".xlsx":
            return pd.read_excel(path, nrows=n), None
        else:
            df = pd.read_parquet(path)
            return df.head(n), None
    except Exception as e:
        return None, str(e)

@st.cache_data(show_spinner=False)
def read_tail(path_str: str, n: int = PREVIEW_N):
    path = Path(path_str)
    try:
        if path.suffix == ".xlsx":
            total = _xlsx_row_count(path)
            skip  = max(1, total - n + 1)  # +1 因为要保留表头(row 0)
            df    = pd.read_excel(path, skiprows=range(1, skip))
            return df, total, None
        else:
            df = pd.read_parquet(path)
            return df.tail(n), len(df), None
    except Exception as e:
        return None, 0, str(e)

@st.cache_data(show_spinner=False)
def read_sample(path_str: str, n: int = SAMPLE_N):
    """读前 n 行用于 profile 统计；返回 (df, total_rows, err)。"""
    path = Path(path_str)
    try:
        if path.suffix == ".xlsx":
            total = _xlsx_row_count(path)
            df    = pd.read_excel(path, nrows=n)
            return df, total, None
        else:
            df = pd.read_parquet(path)
            return df, len(df), None
    except Exception as e:
        return None, 0, str(e)

@st.cache_data(show_spinner=False)
def read_columns(path_str: str):
    """只读列名（用于列匹配），xlsx 只读1行。"""
    path = Path(path_str)
    try:
        if path.suffix == ".xlsx":
            df = pd.read_excel(path, nrows=0)
        else:
            df = pd.read_parquet(path)
        return list(df.columns), None
    except Exception as e:
        return [], str(e)

raw_path     = raw_map[selected_file] if browse_raw else (find_raw_path(file_row, raw_map) if has_raw else None)
cleaned_path = find_cleaned_path(file_row, cleaned_map) if (file_row is not None and has_clean) else None

# ── 列匹配辅助 ───────────────────────────────────────────────────────────────
def get_col_set(cols) -> set[str]:
    return {str(c).strip() for c in cols}

def filter_cols_by_shared(df: pd.DataFrame, ref_cols) -> pd.DataFrame:
    ref_set = get_col_set(ref_cols)
    keep = [c for c in df.columns if str(c).strip() in ref_set]
    return df[keep]

# ── 数据预览辅助（懒读取：前50行 + 后50行）───────────────────────────────────
def show_lazy_preview(path: Path, col_filter=None):
    path_str = str(path)
    with st.spinner("读取前 50 行…"):
        head_df, head_err = read_head(path_str)
    if head_err:
        st.error(f"读取失败：{head_err}")
        return
    if col_filter is not None:
        head_df = filter_cols_by_shared(head_df, col_filter)
    with st.spinner("读取后 50 行…"):
        tail_df, total, tail_err = read_tail(path_str)
    if col_filter is not None and tail_df is not None:
        tail_df = filter_cols_by_shared(tail_df, col_filter)

    row_label = f"{total:,}" if total else "未知"
    st.caption(f"共约 {row_label} 行 × {head_df.shape[1]} 列（大文件仅展示首尾各 {PREVIEW_N} 行）")
    st.markdown(f"**前 {PREVIEW_N} 行**")
    st.dataframe(head_df, use_container_width=True)
    st.divider()
    st.markdown(f"**后 {PREVIEW_N} 行**")
    if tail_err:
        st.error(f"后 {PREVIEW_N} 行读取失败：{tail_err}")
    else:
        st.dataframe(tail_df, use_container_width=True)

def show_lazy_profile(path: Path, label: str):
    path_str = str(path)
    with st.spinner(f"读取 {label} 样本…"):
        sample_df, total, err = read_sample(path_str)
    if err:
        st.error(f"读取失败：{err}")
        return
    is_sample = path.suffix == ".xlsx" and total > SAMPLE_N
    if is_sample:
        st.caption(f"⚠️ 大文件：Profile 基于前 {SAMPLE_N:,} 行样本（总行数约 {total:,}）")
    show_profile(sample_df, label)

# ── 动态 Tab ──────────────────────────────────────────────────────────────────
tab_labels = []
if has_raw:               tab_labels.append("Raw 预览")
if has_clean:             tab_labels.append("Cleaned 预览")
if has_raw:               tab_labels.append("Raw Profile")
if has_clean:             tab_labels.append("Cleaned Profile")
if has_raw and has_clean: tab_labels.append("Profile 对比")

tabs = st.tabs(tab_labels)
idx  = 0

if has_raw:
    with tabs[idx]:
        show_lazy_preview(raw_path)
    idx += 1

if has_clean:
    with tabs[idx]:
        if browse_raw and raw_path is not None:
            raw_cols, _ = read_columns(str(raw_path))
            cln_head, _ = read_head(str(cleaned_path))
            shared_n = len(filter_cols_by_shared(cln_head, raw_cols).columns) if cln_head is not None else 0
            st.caption(f"列名过滤：与 Raw 共同列 {shared_n} 个")
            show_lazy_preview(cleaned_path, col_filter=raw_cols)
        else:
            show_lazy_preview(cleaned_path)
    idx += 1

if has_raw:
    with tabs[idx]:
        show_lazy_profile(raw_path, "Raw")
    idx += 1

if has_clean:
    with tabs[idx]:
        show_lazy_profile(cleaned_path, "Cleaned")
    idx += 1

if has_raw and has_clean:
    with tabs[idx]:
        raw_cols, raw_col_err = read_columns(str(raw_path))
        cln_cols, cln_col_err = read_columns(str(cleaned_path))
        if raw_col_err or cln_col_err:
            if raw_col_err: st.error(f"raw 列名读取失败：{raw_col_err}")
            if cln_col_err: st.error(f"cleaned 列名读取失败：{cln_col_err}")
        else:
            raw_df, raw_total, raw_err2 = read_sample(str(raw_path))
            cln_df, cln_total, cln_err2 = read_sample(str(cleaned_path))
            if raw_err2 or cln_err2:
                if raw_err2: st.error(f"raw 读取失败：{raw_err2}")
                if cln_err2: st.error(f"cleaned 读取失败：{cln_err2}")
            else:
                m1, m2 = st.columns(2)
                m1.metric("Raw 行数（约）",     f"{raw_total:,}")
                m2.metric("Cleaned 行数（约）", f"{cln_total:,}", delta=f"{cln_total-raw_total:+,}", delta_color="normal")
                n1, n2 = st.columns(2)
                n1.metric("Raw 列数",     len(raw_cols))
                n2.metric("Cleaned 列数", len(cln_cols), delta=f"{len(cln_cols)-len(raw_cols):+}", delta_color="normal")
                if raw_total > SAMPLE_N or cln_total > SAMPLE_N:
                    st.caption(f"⚠️ Profile 对比基于各文件前 {SAMPLE_N:,} 行样本")
                st.markdown("#### 逐字段对比")
                show_profile_diff(raw_df, cln_df)