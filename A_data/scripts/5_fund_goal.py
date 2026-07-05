"""根据 iFind“投资目标”文本，为基金目标添加三分类标签。

脚本会读取 ``A_data/data/iFind_terminal_fund_center`` 中所有文件名以
“投资目标.xlsx”结尾的工作簿，并把结果保存到
``A_data/prepared_data/iFind_terminal_fund_center``。输出文件与源文件同名，
原有工作表、单元格格式和冻结窗格会尽量保留，只在含“投资目标”列的工作表末尾
增加两列：

1. ``投资目标分类代码``：1、2 或 3；
2. ``投资目标分类``：绝对收益类、相对市场类或相对基准指数类。

分类规则来自 ``A_data/reference/fund_goal.json``。规则按优先级依次匹配：
明确提到业绩基准或指数的文本优先归为 3；提到市场比较或超额收益的文本归为 2；
其余非空投资目标归为 1。空白文本不强行推断，两个新增字段均保留为空。

运行方式：

    .venv/bin/python A_data/scripts/5_fund_goal.py
"""

from __future__ import annotations

import argparse
import json
import re
import tempfile
from collections import Counter
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
A_DATA_ROOT = PROJECT_ROOT / "A_data"

DEFAULT_INPUT_DIR = A_DATA_ROOT / "data" / "iFind_terminal_fund_center"
DEFAULT_OUTPUT_DIR = A_DATA_ROOT / "prepared_data" / "iFind_terminal_fund_center"
DEFAULT_RULES_PATH = A_DATA_ROOT / "reference" / "fund_goal.json"

SOURCE_COLUMN = "投资目标"
CODE_COLUMN = "投资目标分类代码"
LABEL_COLUMN = "投资目标分类"


@dataclass(frozen=True)
class CategoryRule:
    """保存一个已经通过校验并完成正则编译的分类规则。"""

    code: int
    label: str
    priority: int
    patterns: tuple[re.Pattern[str], ...]
    is_default: bool = False


def parse_args() -> argparse.Namespace:
    """读取命令行参数，方便将来对其他目录或规则文件复用本脚本。"""
    parser = argparse.ArgumentParser(
        description="根据‘投资目标’文本为 iFind 基金工作簿增加目标分类。"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="以‘投资目标.xlsx’结尾的源工作簿所在目录。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="分类后工作簿的输出目录。",
    )
    parser.add_argument(
        "--rules",
        type=Path,
        default=DEFAULT_RULES_PATH,
        help="正则分类规则 JSON 文件。",
    )
    return parser.parse_args()


def _require_mapping(value: Any, field_name: str) -> dict[str, Any]:
    """确保 JSON 中应为对象的字段确实是对象，并给出清楚的报错。"""
    if not isinstance(value, dict):
        raise ValueError(f"规则字段 {field_name!r} 必须是 JSON 对象。")
    return value


def _require_list(value: Any, field_name: str) -> list[Any]:
    """确保 JSON 中应为数组的字段确实是数组。"""
    if not isinstance(value, list):
        raise ValueError(f"规则字段 {field_name!r} 必须是 JSON 数组。")
    return value


def load_rules(path: Path) -> tuple[list[CategoryRule], dict[str, Any]]:
    """读取并校验规则文件，同时提前编译正则以尽早发现拼写错误。"""
    if not path.exists():
        raise FileNotFoundError(f"找不到分类规则文件：{path}")

    with path.open("r", encoding="utf-8") as file:
        config = _require_mapping(json.load(file), "根节点")

    normalization = _require_mapping(
        config.get("normalization", {}), "normalization"
    )
    category_configs = _require_list(config.get("categories"), "categories")
    if not category_configs:
        raise ValueError("规则文件 categories 不能为空。")

    rules: list[CategoryRule] = []
    seen_codes: set[int] = set()
    default_count = 0

    for index, raw_rule in enumerate(category_configs):
        rule = _require_mapping(raw_rule, f"categories[{index}]")
        try:
            code = int(rule["code"])
            label = str(rule["label"]).strip()
            priority = int(rule["priority"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError(
                f"categories[{index}] 的 code、label 或 priority 无效。"
            ) from error

        if code not in {1, 2, 3}:
            raise ValueError(f"分类代码只能是 1、2、3，当前值为 {code}。")
        if code in seen_codes:
            raise ValueError(f"分类代码 {code} 在规则文件中重复。")
        if not label:
            raise ValueError(f"分类代码 {code} 的 label 不能为空。")
        seen_codes.add(code)

        is_default = bool(rule.get("default", False))
        default_count += int(is_default)
        raw_patterns = _require_list(rule.get("match_any", []), "match_any")

        compiled_patterns: list[re.Pattern[str]] = []
        for pattern_index, raw_pattern in enumerate(raw_patterns):
            if not isinstance(raw_pattern, str) or not raw_pattern:
                raise ValueError(
                    f"分类代码 {code} 的第 {pattern_index + 1} 个正则必须是非空字符串。"
                )
            try:
                compiled_patterns.append(re.compile(raw_pattern))
            except re.error as error:
                raise ValueError(
                    f"分类代码 {code} 的正则无法编译：{raw_pattern!r}"
                ) from error

        rules.append(
            CategoryRule(
                code=code,
                label=label,
                priority=priority,
                patterns=tuple(compiled_patterns),
                is_default=is_default,
            )
        )

    if seen_codes != {1, 2, 3}:
        raise ValueError("规则文件必须且只能定义分类代码 1、2、3。")
    if default_count != 1:
        raise ValueError("规则文件必须且只能设置一个 default 分类。")

    # 优先级较小的规则先匹配。这样即使文本同时出现“超额收益”和“业绩基准”，
    # 也会先识别更明确的业绩基准，避免被较宽泛的“超额收益”截走。
    rules.sort(key=lambda item: item.priority)
    return rules, normalization


def normalize_goal(value: object, normalization: dict[str, Any]) -> str:
    """统一投资目标文本，减少换行、全角空格和导出标记对正则的干扰。"""
    if value is None or pd.isna(value):
        return ""

    text = str(value)
    for literal in normalization.get("remove_literals", []):
        text = text.replace(str(literal), "")
    if normalization.get("remove_whitespace", False):
        text = re.sub(r"\s+", "", text)
    return text.strip()


def classify_goal(
    value: object,
    rules: Iterable[CategoryRule],
    normalization: dict[str, Any],
) -> tuple[int | None, str | None]:
    """按规则优先级分类；空文本返回空值，不用默认类别掩盖缺失数据。"""
    text = normalize_goal(value, normalization)
    if not text:
        return None, None

    default_rule: CategoryRule | None = None
    for rule in rules:
        if rule.is_default:
            default_rule = rule
            continue
        if any(pattern.search(text) for pattern in rule.patterns):
            return rule.code, rule.label

    if default_rule is None:
        raise RuntimeError("分类规则缺少 default 类别。")
    return default_rule.code, default_rule.label


def discover_source_files(input_dir: Path) -> list[Path]:
    """寻找源文件，并避免把 Excel 临时锁文件当作正式输入。"""
    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在：{input_dir}")

    files = sorted(
        path
        for path in input_dir.glob("*投资目标.xlsx")
        if not path.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(
            f"{input_dir} 中没有找到以‘投资目标.xlsx’结尾的文件。"
        )
    return files


def copy_cell_style(source: Cell, target: Cell) -> None:
    """把相邻单元格的显示格式复制给新增列，但不复制原来的值。"""
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def classify_sheet(
    worksheet: Any,
    frame: pd.DataFrame,
    rules: list[CategoryRule],
    normalization: dict[str, Any],
) -> Counter[str]:
    """分类一个工作表，并把结果写入 openpyxl 工作表对象的末尾两列。"""
    cleaned_headers = [str(column).replace("↑", "").replace("↓", "").strip() for column in frame.columns]
    if SOURCE_COLUMN not in cleaned_headers:
        raise ValueError(
            f"工作表 {worksheet.title!r} 缺少‘{SOURCE_COLUMN}’列；"
            f"实际列名为：{cleaned_headers}"
        )

    source_index = cleaned_headers.index(SOURCE_COLUMN)
    source_series = frame.iloc[:, source_index]
    classifications = [
        classify_goal(value, rules, normalization) for value in source_series
    ]

    # 每次都根据原始表头定位新增列。若输出文件此前已经存在，脚本仍从源文件重建，
    # 因此不会在重复运行时不断叠加同名分类列。
    code_column_index = worksheet.max_column + 1
    label_column_index = worksheet.max_column + 2
    style_source_column = max(1, code_column_index - 1)

    code_header = worksheet.cell(row=1, column=code_column_index, value=CODE_COLUMN)
    label_header = worksheet.cell(row=1, column=label_column_index, value=LABEL_COLUMN)
    copy_cell_style(worksheet.cell(row=1, column=style_source_column), code_header)
    copy_cell_style(worksheet.cell(row=1, column=style_source_column), label_header)

    counts: Counter[str] = Counter()
    for excel_row, (code, label) in enumerate(classifications, start=2):
        code_cell = worksheet.cell(row=excel_row, column=code_column_index, value=code)
        label_cell = worksheet.cell(row=excel_row, column=label_column_index, value=label)
        copy_cell_style(
            worksheet.cell(row=excel_row, column=style_source_column), code_cell
        )
        copy_cell_style(
            worksheet.cell(row=excel_row, column=style_source_column), label_cell
        )
        counts[label if label is not None else "未分类（投资目标为空）"] += 1

    # 新增字段比原有短字段更长，给它们明确且适中的列宽，避免表头被截断。
    worksheet.column_dimensions[get_column_letter(code_column_index)].width = 18
    worksheet.column_dimensions[get_column_letter(label_column_index)].width = 20
    return counts


def process_workbook(
    source_path: Path,
    output_path: Path,
    rules: list[CategoryRule],
    normalization: dict[str, Any],
) -> Counter[str]:
    """处理一个工作簿，并通过临时文件原子替换输出，避免留下半写入文件。"""
    workbook = load_workbook(source_path, data_only=False, keep_links=True)
    workbook_counts: Counter[str] = Counter()
    processed_sheets = 0

    # pandas 负责稳定地读取单元格值，openpyxl 负责在原工作簿上追加字段，
    # 两者配合可以避免重新导出时丢失原有格式、冻结窗格或工作表属性。
    frames = pd.read_excel(source_path, sheet_name=None)
    for sheet_name, frame in frames.items():
        cleaned_headers = {
            str(column).replace("↑", "").replace("↓", "").strip()
            for column in frame.columns
        }
        if SOURCE_COLUMN not in cleaned_headers:
            continue

        worksheet = workbook[sheet_name]
        if len(frame) != worksheet.max_row - 1:
            raise ValueError(
                f"{source_path.name} / {sheet_name} 的 pandas 数据行数 {len(frame)} "
                f"与 Excel 行数 {worksheet.max_row - 1} 不一致，无法安全写回。"
            )
        workbook_counts.update(
            classify_sheet(worksheet, frame, rules, normalization)
        )
        processed_sheets += 1

    if processed_sheets == 0:
        raise ValueError(f"{source_path} 中没有含‘{SOURCE_COLUMN}’列的工作表。")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=output_path.parent,
            prefix=f".{output_path.stem}_",
            suffix=".xlsx",
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()

    return workbook_counts


def main() -> None:
    """批量分类全部目标工作簿，并打印便于核对的分类统计。"""
    args = parse_args()
    rules, normalization = load_rules(args.rules)
    source_files = discover_source_files(args.input_dir)

    total_counts: Counter[str] = Counter()
    for source_path in source_files:
        output_path = args.output_dir / source_path.name
        counts = process_workbook(
            source_path=source_path,
            output_path=output_path,
            rules=rules,
            normalization=normalization,
        )
        total_counts.update(counts)
        count_text = "，".join(
            f"{label} {count} 行" for label, count in sorted(counts.items())
        )
        print(f"完成：{source_path.name} -> {output_path}（{count_text}）")

    print(
        "合计："
        + "，".join(
            f"{label} {count} 行" for label, count in sorted(total_counts.items())
        )
    )


if __name__ == "__main__":
    main()
