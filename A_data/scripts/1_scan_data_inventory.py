"""
1_scan_data_inventory.py

用途：
固定扫描 A_data/data 文件夹（不含子文件夹 old_data_archive），
提取每个 .xlsx / .parquet 文件的基础结构信息，
生成 A_data/output/data_inventory.xlsx。

支持 .scanignore 文件（放在 A_data/data/ 下），效果类似 .gitignore。

注意：
- old_data_archive 子文件夹始终被跳过，无需写入 .scanignore。
- 只扫描 .xlsx（不含 .xls）和 .parquet。
- 自动跳过 Excel 临时文件，例如 ~$xxx.xlsx。
"""

from pathlib import Path
from collections import deque
from fnmatch import fnmatch
import json

import pandas as pd
from openpyxl import load_workbook

ALWAYS_IGNORE_DIRS = {"old_data_archive"}


IGNORE_FILE_NAME = ".scanignore"


def to_posix_path(path: Path) -> str:
    """
    把 Path 转成统一的正斜杠格式，避免 Windows 反斜杠造成匹配问题。
    """
    return str(path).replace("\\", "/")


def load_scanignore_patterns(root_dir: Path) -> list[str]:
    """
    读取 data_public/.scanignore 中的忽略规则。

    返回：
    patterns: 忽略规则列表
    """
    ignore_file = root_dir / IGNORE_FILE_NAME

    if not ignore_file.exists():
        return []

    patterns = []

    with open(ignore_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            # 跳过空行和注释
            if not line:
                continue
            if line.startswith("#"):
                continue

            # 统一路径分隔符
            line = line.replace("\\", "/")
            patterns.append(line)

    return patterns


def should_ignore_path(file_path: Path, root_dir: Path, ignore_patterns: list[str]) -> bool:
    """
    判断某个文件是否应该被忽略。

    支持：
    1. 普通忽略规则：
        宏观大市/*.xlsx
        daily/MKTP/
        *测试*.xlsx

    2. 反向保留规则：
        !宏观大市/micro_combination.xlsx

    规则说明：
    - 后面的规则可以覆盖前面的规则。
    - 如果先写 宏观大市/*.xlsx，再写 !宏观大市/micro_combination.xlsx，
      则表示忽略 宏观大市 下所有 xlsx，但保留 micro_combination.xlsx。
    """
    relative_path = file_path.relative_to(root_dir)
    relative_posix = to_posix_path(relative_path)
    file_name = file_path.name

    ignored = False

    for raw_pattern in ignore_patterns:
        pattern = raw_pattern.strip()

        if not pattern:
            continue

        is_negative_rule = pattern.startswith("!")

        if is_negative_rule:
            pattern = pattern[1:].strip()

        matched = False

        # 情况 1：文件夹规则
        # 例如：
        # daily/MKTP/
        # backup/
        if pattern.endswith("/"):
            folder_pattern = pattern.rstrip("/")

            if relative_posix == folder_pattern:
                matched = True

            if relative_posix.startswith(folder_pattern + "/"):
                matched = True

            path_parts = relative_posix.split("/")
            if folder_pattern in path_parts:
                matched = True

        # 情况 2：完整相对路径规则
        # 例如：
        # 宏观大市/*.xlsx
        # 宏观大市/micro_combination.xlsx
        if fnmatch(relative_posix, pattern):
            matched = True

        # 情况 3：文件名规则
        # 例如：
        # *测试*.xlsx
        if fnmatch(file_name, pattern):
            matched = True

        if matched:
            if is_negative_rule:
                ignored = False
            else:
                ignored = True

    return ignored


def to_json_safe(value):
    """
    把 Python 对象转成适合写入 Excel 的 JSON 字符串。
    """
    return json.dumps(value, ensure_ascii=False, default=str)


def is_non_numeric_value(x) -> bool:
    """
    判断单个值是否属于“非空且无法转成数字”。

    规则：
    - 空值 / None / 空字符串：不算非数字
    - 可以被 pd.to_numeric 转换：算数字
    - 转换失败：算非数字
    """
    if x is None:
        return False

    if isinstance(x, str) and x.strip() == "":
        return False

    converted = pd.to_numeric(pd.Series([x]), errors="coerce").iloc[0]

    return pd.isna(converted)


def row_has_non_numeric_after_first_col(row_values) -> bool:
    """
    判断一整行在“去掉首列后”是否存在非数字内容。

    参数：
    row_values: list，例如 Excel 某一行的所有单元格值

    返回：
    True  = 这一行除首列外，存在非空且无法转成数字的内容
    False = 这一行除首列外，不存在这类内容
    """
    values_after_first_col = row_values[1:]

    for value in values_after_first_col:
        if is_non_numeric_value(value):
            return True

    return False


def scan_one_sheet(ws):
    """
    扫描一个 worksheet。

    返回：
    - first_row_values: Excel 第一行内容
    - head5_rows: 数据区前五行，默认从 Excel 第 2 行开始，因为第 1 行通常是列名
    - tail5_rows: 数据区最后五行
    - head_non_numeric_pos: 前五行中，哪些样本内行号存在非数字内容，例如 [0, 1, 2]
    - head_non_numeric_excel_row: 对应 Excel 真实行号
    - tail_non_numeric_pos
    - tail_non_numeric_excel_row
    """
    first_row_values = []
    head5_rows = []
    tail5_deque = deque(maxlen=5)

    for excel_row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = list(row)

        # 第一行：一般是列名
        if excel_row_num == 1:
            first_row_values = row_values
            continue

        # 跳过完全空行，避免尾部格式行干扰
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
            continue

        row_record = {
            "excel_row": excel_row_num,
            "values": row_values,
        }

        # 数据区前五行
        if len(head5_rows) < 5:
            head5_rows.append(row_record)

        # 数据区后五行
        tail5_deque.append(row_record)

    tail5_rows = list(tail5_deque)

    head_non_numeric_pos = []
    head_non_numeric_excel_row = []

    for pos, row_record in enumerate(head5_rows):
        if row_has_non_numeric_after_first_col(row_record["values"]):
            head_non_numeric_pos.append(pos)
            head_non_numeric_excel_row.append(row_record["excel_row"])

    tail_non_numeric_pos = []
    tail_non_numeric_excel_row = []

    for pos, row_record in enumerate(tail5_rows):
        if row_has_non_numeric_after_first_col(row_record["values"]):
            tail_non_numeric_pos.append(pos)
            tail_non_numeric_excel_row.append(row_record["excel_row"])

    return {
        "first_row_values": first_row_values,
        "head5_rows": head5_rows,
        "tail5_rows": tail5_rows,
        "head_non_numeric_pos": head_non_numeric_pos,
        "head_non_numeric_excel_row": head_non_numeric_excel_row,
        "tail_non_numeric_pos": tail_non_numeric_pos,
        "tail_non_numeric_excel_row": tail_non_numeric_excel_row,
    }


def scan_parquet_file(file_path: Path) -> dict:
    """
    扫描一个 parquet 文件，返回结构信息。
    """
    df = pd.read_parquet(file_path)
    columns = list(df.columns)
    head5 = df.head(5).values.tolist()
    tail5 = df.tail(5).values.tolist()
    return {
        "columns": columns,
        "row_count": len(df),
        "col_count": len(columns),
        "head5": head5,
        "tail5": tail5,
    }


def scan_all_files(root_dir: Path) -> pd.DataFrame:
    """
    扫描 root_dir 及其子文件夹下所有 .xlsx 和 .parquet 文件，
    自动跳过 ALWAYS_IGNORE_DIRS 中的子文件夹。

    参数：
    root_dir: A_data/data 文件夹路径

    返回：
    inventory_df: 每个文件（xlsx 每 sheet 一行，parquet 每文件一行）的盘点表
    """
    records = []

    ignore_patterns = load_scanignore_patterns(root_dir)

    all_files = sorted(
        p for p in root_dir.rglob("*")
        if p.suffix in {".xlsx", ".parquet"} and p.is_file()
    )

    for file_path in all_files:
        # 跳过 ALWAYS_IGNORE_DIRS 中的子文件夹
        relative_parts = file_path.relative_to(root_dir).parts
        if any(part in ALWAYS_IGNORE_DIRS for part in relative_parts):
            continue

        # 跳过 Excel 临时文件
        if file_path.name.startswith("~$"):
            continue

        # 跳过结果文件本身
        if file_path.name == "data_inventory.xlsx":
            continue

        # 根据 .scanignore 判断是否跳过
        if should_ignore_path(file_path, root_dir, ignore_patterns):
            continue

        relative_path = Path(root_dir.name) / file_path.relative_to(root_dir)

        print(f"正在扫描：{to_posix_path(relative_path)}")

        relative_parts = file_path.relative_to(root_dir).parts
        if "data_csmar" in relative_parts:
            data_source = "data_csmar"
        else:
            data_source = file_path.parent.name

        # ── parquet ──────────────────────────────────────────────
        if file_path.suffix == ".parquet":
            try:
                info = scan_parquet_file(file_path)
                records.append({
                    "file_type": "parquet",
                    "data_source": data_source,
                    "file_name": file_path.name,
                    "file_path": to_posix_path(relative_path),
                    "sheet_name": "",
                    "max_row": info["row_count"],
                    "max_column": info["col_count"],
                    "first_row_values": to_json_safe(info["columns"]),
                    "head5_values": to_json_safe(info["head5"]),
                    "tail5_values": to_json_safe(info["tail5"]),
                    "head_non_numeric_pos": "",
                    "head_non_numeric_excel_row": "",
                    "tail_non_numeric_pos": "",
                    "tail_non_numeric_excel_row": "",
                    "error": "",
                })
            except Exception as e:
                records.append({
                    "file_type": "parquet",
                    "data_source": data_source,
                    "file_name": file_path.name,
                    "file_path": to_posix_path(relative_path),
                    "sheet_name": "", "max_row": "", "max_column": "",
                    "first_row_values": "", "head5_values": "", "tail5_values": "",
                    "head_non_numeric_pos": "", "head_non_numeric_excel_row": "",
                    "tail_non_numeric_pos": "", "tail_non_numeric_excel_row": "",
                    "error": f"file_error: {e}",
                })
            continue

        # ── xlsx ─────────────────────────────────────────────────
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                try:
                    sheet_info = scan_one_sheet(ws)
                    records.append({
                        "file_type": "xlsx",
                        "data_source": data_source,
                        "file_name": file_path.name,
                        "file_path": to_posix_path(relative_path),
                        "sheet_name": sheet_name,
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                        "first_row_values": to_json_safe(sheet_info["first_row_values"]),
                        "head5_values": to_json_safe(sheet_info["head5_rows"]),
                        "tail5_values": to_json_safe(sheet_info["tail5_rows"]),
                        "head_non_numeric_pos": to_json_safe(sheet_info["head_non_numeric_pos"]),
                        "head_non_numeric_excel_row": to_json_safe(sheet_info["head_non_numeric_excel_row"]),
                        "tail_non_numeric_pos": to_json_safe(sheet_info["tail_non_numeric_pos"]),
                        "tail_non_numeric_excel_row": to_json_safe(sheet_info["tail_non_numeric_excel_row"]),
                        "error": "",
                    })
                except Exception as sheet_error:
                    records.append({
                        "file_type": "xlsx",
                        "data_source": data_source,
                        "file_name": file_path.name,
                        "file_path": to_posix_path(relative_path),
                        "sheet_name": sheet_name,
                        "max_row": "", "max_column": "",
                        "first_row_values": "", "head5_values": "", "tail5_values": "",
                        "head_non_numeric_pos": "", "head_non_numeric_excel_row": "",
                        "tail_non_numeric_pos": "", "tail_non_numeric_excel_row": "",
                        "error": f"sheet_error: {sheet_error}",
                    })

            wb.close()

        except Exception as file_error:
            records.append({
                "file_type": "xlsx",
                "data_source": data_source,
                "file_name": file_path.name,
                "file_path": to_posix_path(relative_path),
                "sheet_name": "", "max_row": "", "max_column": "",
                "first_row_values": "", "head5_values": "", "tail5_values": "",
                "head_non_numeric_pos": "", "head_non_numeric_excel_row": "",
                "tail_non_numeric_pos": "", "tail_non_numeric_excel_row": "",
                "error": f"file_error: {file_error}",
            })

    return pd.DataFrame(records)


def main():
    """
    主程序。固定扫描 A_data/data（排除 old_data_archive），
    输出到 A_data/reference/data_inventory_A.xlsx 并转换为 JSON。
    """
    import subprocess

    script_dir = Path(__file__).resolve().parent        # A_data/scripts
    root_dir = script_dir.parent / "data"               # A_data/data
    ref_dir = script_dir.parent / "reference"           # A_data/reference

    if not root_dir.exists():
        print(f"错误：找不到扫描目录 {root_dir}")
        return

    ref_dir.mkdir(parents=True, exist_ok=True)

    ignore_patterns = load_scanignore_patterns(root_dir)

    if ignore_patterns:
        print("已读取 .scanignore，忽略规则如下：")
        for pattern in ignore_patterns:
            print(f"  - {pattern}")
    else:
        print("未发现 .scanignore，扫描范围：A_data/data（排除 old_data_archive）。")

    inventory_df = scan_all_files(root_dir)

    # 重排列順序：前四列為手動填寫欄位，file_type 移至最後
    MANUAL_COLS = ["clean_data", "frequency", "content"]
    inventory_df.insert(0, "clean_data", "")
    inventory_df.insert(1, "frequency", "")
    inventory_df.insert(2, "content", "")

    xlsx_output = ref_dir / "data_inventory_A.xlsx"

    # 合并旧文件中的手填列
    if xlsx_output.exists():
        try:
            old_df = pd.read_excel(xlsx_output)
            merge_key = ["file_path", "sheet_name"]
            if all(c in old_df.columns for c in merge_key + MANUAL_COLS):
                old_manual = old_df[merge_key + MANUAL_COLS].copy()
                old_manual = old_manual.fillna("")
                inventory_df = inventory_df.drop(columns=MANUAL_COLS)
                inventory_df = inventory_df.merge(old_manual, on=merge_key, how="left")
                inventory_df[MANUAL_COLS] = inventory_df[MANUAL_COLS].fillna("")
                print(f"已从旧文件合并手填列（{len(old_manual)} 条旧记录）。")
        except Exception as e:
            print(f"警告：读取旧文件失败，手填列将为空。({e})")

    fixed_cols = MANUAL_COLS + ["data_source"]
    remaining = [c for c in inventory_df.columns if c not in fixed_cols and c != "file_type"]
    inventory_df = inventory_df[fixed_cols + remaining + ["file_type"]]

    inventory_df.to_excel(xlsx_output, index=False)
    print(f"扫描完成，共记录 {len(inventory_df)} 个条目（xlsx sheet + parquet 文件）。")
    print(f"xlsx 已保存到：{xlsx_output}")

    # ── 转换为 JSON ──────────────────────────────────────────────
    converter = Path.home() / "Projects/xlsx_to_jsonl_project/scripts/convert_xlsx_to_jsonl.py"
    json_output = ref_dir / "data_inventory_A.json"

    print(f"\n正在转换为 JSON ...")
    result = subprocess.run(
        [
            "python", str(converter),
            "--input", str(xlsx_output),
            "--output", str(json_output),
        ],
        capture_output=True,
        text=True,
    )

    if result.returncode == 0:
        print(result.stdout.strip())
        print(f"json 已保存到：{json_output}")
    else:
        print(f"警告：JSON 转换失败。\n{result.stderr.strip()}")


if __name__ == "__main__":
    main()