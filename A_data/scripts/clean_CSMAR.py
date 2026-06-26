#!/usr/bin/env python3
"""Interactively clean CSMAR three-row-header Excel files.

The script reads one or more CSMAR workbooks, validates that all selected
sheets share the same three header rows, cleans and merges the data rows, and
writes the result plus a small inventory under A_data/output.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
A_DATA_ROOT = PROJECT_ROOT / "A_data"
DATA_DIR = A_DATA_ROOT / "data"
OUTPUT_DIR = A_DATA_ROOT / "output"
INVENTORY_PATH = A_DATA_ROOT / "reference" / "data_inventory_A.json"
DATA_GUIDE_PATH = OUTPUT_DIR / "data_guide.md"
CONVERT_SCRIPT = PROJECT_ROOT.parent / "jsonl_to_xlsx_project" / "scripts" / "convert_jsonl_to_xlsx.py"

NA_TEXT = {"", "nan", "none", "null", "na", "n/a", "--", "-", "没有单位"}
DATE_NAME_RE = re.compile(r"(date|accper|startdate|enddate|日期|时间|月份|季度|年度)", re.IGNORECASE)
TIME_SERIES_DATE_RE = re.compile(
    r"^(date|tradingdate|accper|enddate|statdate|startdate|month|quarter|year|交易日期|日期|统计日期|截止日期)$",
    re.IGNORECASE,
)
STRING_NAME_RE = re.compile(
    r"(code|symbol|id|name|type|status|category|style|source|flag|"
    r"基金|证券|代码|名称|类型|类别|状态|标识|说明|公司|机构|目标|范围|策略|律师|会计|托管)",
    re.IGNORECASE,
)
SPLIT_SUFFIX_RE = re.compile(r"\d+$")


@dataclass
class SheetPayload:
    source_path: Path
    sheet_name: str
    raw_header: list[list[str]]
    first_raw_values: list[str]
    frame: pd.DataFrame
    date_col: str | None
    frequency: str


@dataclass
class SheetMeta:
    source_path: Path
    sheet_name: str
    raw_header: list[list[str]]
    first_raw_values: list[str]
    date_col: str | None
    frequency: str = "uncertain"


def project_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path.resolve())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "files",
        nargs="*",
        help="Input Excel paths. Multiple paths may also be supplied with --files.",
    )
    parser.add_argument(
        "--files",
        dest="files_option",
        action="append",
        default=[],
        help="Comma-separated input Excel paths. May be repeated.",
    )
    parser.add_argument("--output", help="Output file name. Defaults to an interactive prompt.")
    parser.add_argument("--content", help="Content description. Defaults to an interactive prompt.")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    return parser.parse_args()


def split_paths(values: list[str]) -> list[str]:
    paths: list[str] = []
    for value in values:
        for part in re.split(r"[,，\n]+", value):
            cleaned = part.strip().strip("\"'")
            if cleaned:
                paths.append(cleaned)
    return paths


def prompt_for_paths() -> list[str]:
    print("请输入需要清洗的文件路径；可用逗号分隔，或每行输入一个路径。输入空行结束：")
    lines = []
    while True:
        line = input("> ").strip()
        if not line:
            break
        lines.append(line)
    return split_paths(lines)


def resolve_input_path(value: str) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def normalize_output_name(name: str) -> str:
    output = Path(name.strip().strip("\"'")).name
    if not output:
        raise SystemExit("Output file name cannot be empty.")
    if not Path(output).suffix:
        output = f"{output}.parquet"
    return output


def normalize_cell(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in NA_TEXT else text


def clean_scalar(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, str):
        text = value.strip()
        return pd.NA if text.lower() in NA_TEXT else text
    return value


def unique_column_names(values: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    columns = []
    for idx, value in enumerate(values):
        base = value or f"col_{idx}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        columns.append(base if count == 0 else f"{base}_{count + 1}")
    return columns


def read_sheet(path: Path, sheet_name: str) -> tuple[list[list[str]], list[str], pd.DataFrame]:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
    raw = raw.dropna(how="all")
    if raw.shape[0] < 4:
        raise ValueError(f"{project_path(path)}::{sheet_name} has fewer than four non-empty rows")

    header = [[normalize_cell(v) for v in raw.iloc[row].tolist()] for row in range(3)]
    columns = unique_column_names(header[0])
    frame = raw.iloc[3:].copy()
    frame.columns = columns
    frame = frame.dropna(how="all").reset_index(drop=True)
    return header, header[0], frame


def date_candidate_score(column: str, series: pd.Series) -> tuple[int, int]:
    name = str(column).strip()
    if not DATE_NAME_RE.search(name):
        return (0, 0)
    parsed = pd.to_datetime(series, errors="coerce")
    valid = int(parsed.notna().sum())
    if valid == 0:
        return (0, 0)
    score = valid
    if TIME_SERIES_DATE_RE.search(name):
        score += 1_000_000
    elif DATE_NAME_RE.search(name):
        score += 100_000
    return (score, valid)


def find_date_col(frame: pd.DataFrame) -> str | None:
    candidates = [
        (date_candidate_score(str(column), frame[column]), str(column))
        for column in frame.columns
    ]
    candidates = [item for item in candidates if item[0][1] > 0]
    if not candidates:
        return None
    candidates.sort(reverse=True)
    best_score, best_col = candidates[0]
    if best_score[0] < 1_000_000:
        return None
    return best_col


def infer_frequency(dates: pd.Series | pd.Index | None) -> str:
    if dates is None:
        return "unknown"
    parsed = pd.Series(pd.to_datetime(dates, errors="coerce")).dropna().drop_duplicates().sort_values()
    if len(parsed) < 4:
        return "unknown"

    # Infer from which months appear in the data
    unique_months = set(parsed.dt.month.unique())
    if unique_months <= {12}:
        return "yearly"
    if unique_months <= {6, 12}:
        return "semiannually"
    quarterly_set = {3, 6, 9, 12, 1}
    if unique_months <= quarterly_set and len(unique_months & quarterly_set) >= 4:
        return "quarterly"

    # Infer from day-level diffs
    diffs = parsed.diff().dropna().dt.days
    if not diffs.empty:
        pct_1 = float((diffs == 1).mean())
        pct_3 = float((diffs == 3).mean())
        if pct_1 >= 0.6 and (pct_1 + pct_3) >= 0.9:
            return "trading_day"
        if pct_1 >= 0.8:
            return "daily"

    # Infer from month-level diffs
    periods = parsed.dt.to_period("M")
    month_diffs = periods.astype("int64").diff().dropna()
    if not month_diffs.empty:
        pct_month_1 = float((month_diffs == 1).mean())
        if pct_month_1 >= 0.8:
            return "monthly"

    return "unknown"


def clean_frame(frame: pd.DataFrame) -> tuple[pd.DataFrame, str | None, str]:
    frame = frame.copy()
    for column in frame.columns:
        frame[column] = frame[column].map(clean_scalar)

    date_col = find_date_col(frame)
    for column in frame.columns:
        name = str(column)
        if DATE_NAME_RE.search(name):
            parsed = pd.to_datetime(frame[column], errors="coerce")
            if parsed.notna().any():
                frame[column] = parsed
                continue
        if STRING_NAME_RE.search(name):
            frame[column] = frame[column].where(frame[column].notna(), pd.NA).astype("string").str.strip()
        else:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")

    if "Typrep" in frame.columns:
        before = len(frame)
        frame = frame.loc[frame["Typrep"].astype("string").str.strip().eq("A")].copy()
        if before and frame.empty:
            print("Warning: Typrep exists but no Typrep == 'A' rows remain after filtering.")

    frequency = infer_frequency(frame[date_col]) if date_col and date_col in frame.columns else "uncertain"
    return frame.reset_index(drop=True), date_col, frequency


def dedupe_frame(frame: pd.DataFrame, date_col: str | None) -> pd.DataFrame:
    if frame.empty:
        return frame
    if not date_col:
        return frame
    key_cols = []
    for column in frame.columns:
        name = str(column)
        if column == date_col or re.search(r"(code|symbol|id|代码|证券)", name, re.IGNORECASE):
            key_cols.append(column)
    key_cols = list(dict.fromkeys(key_cols))
    if not key_cols:
        return frame
    return frame.sort_values(by=key_cols, na_position="last").drop_duplicates(subset=key_cols, keep="last")


def load_payloads(paths: list[Path]) -> list[SheetPayload]:
    payloads: list[SheetPayload] = []
    for path in paths:
        if not path.exists():
            raise SystemExit(f"Input file not found: {project_path(path)}")
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise SystemExit(f"Unsupported input file type: {project_path(path)}")
        workbook = pd.ExcelFile(path)
        for sheet_name in workbook.sheet_names:
            header, first_raw_values, raw_frame = read_sheet(path, sheet_name)
            frame, date_col, frequency = clean_frame(raw_frame)
            payloads.append(
                SheetPayload(
                    source_path=path,
                    sheet_name=sheet_name,
                    raw_header=header,
                    first_raw_values=first_raw_values,
                    frame=frame,
                    date_col=date_col,
                    frequency=frequency,
                )
            )
    return payloads


def validate_headers(payloads: list[SheetPayload]) -> None:
    if not payloads:
        raise SystemExit("No sheets found to clean.")
    expected = payloads[0].raw_header
    mismatches = [
        f"{project_path(payload.source_path)}::{payload.sheet_name}"
        for payload in payloads[1:]
        if payload.raw_header != expected
    ]
    if mismatches:
        details = "\n  ".join(mismatches)
        raise SystemExit(
            "The first three header rows are not identical across all selected sheets. "
            "No output was written.\nMismatched sheets:\n  "
            + details
        )


def combine_payloads(payloads: list[SheetPayload]) -> tuple[pd.DataFrame, str | None, str]:
    frames = [payload.frame for payload in payloads if not payload.frame.empty]
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=payloads[0].frame.columns)

    date_cols = [payload.date_col for payload in payloads if payload.date_col]
    date_col = date_cols[0] if date_cols and all(col == date_cols[0] for col in date_cols) else None
    combined = dedupe_frame(combined, date_col)
    if date_col and date_col in combined.columns:
        combined[date_col] = pd.to_datetime(combined[date_col], errors="coerce")
        combined = combined.sort_values(by=date_col, na_position="last")
        frequency = infer_frequency(combined[date_col])
    else:
        frequency = "uncertain"
    return combined.reset_index(drop=True), date_col, frequency


def read_header_stream(path: Path, sheet_name: str) -> tuple[list[list[str]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        sheet.reset_dimensions()
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
            rows.append([normalize_cell(value) for value in row])
        if len(rows) < 3:
            raise ValueError(f"{project_path(path)}::{sheet_name} has fewer than three header rows")
        max_len = max(len(row) for row in rows)
        header = [row + [""] * (max_len - len(row)) for row in rows]
        return header, header[0]
    finally:
        workbook.close()


def collect_sheet_metas(paths: list[Path]) -> list[SheetMeta]:
    metas: list[SheetMeta] = []
    for path in paths:
        if not path.exists():
            raise SystemExit(f"Input file not found: {project_path(path)}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise SystemExit(f"Streaming parquet output supports .xlsx/.xlsm inputs: {project_path(path)}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        for sheet_name in sheet_names:
            header, first_raw_values = read_header_stream(path, sheet_name)
            columns = unique_column_names(first_raw_values)
            date_col = next((column for column in columns if TIME_SERIES_DATE_RE.search(str(column))), None)
            metas.append(
                SheetMeta(
                    source_path=path,
                    sheet_name=sheet_name,
                    raw_header=header,
                    first_raw_values=first_raw_values,
                    date_col=date_col,
                )
            )
    return metas


def validate_meta_headers(metas: list[SheetMeta]) -> None:
    if not metas:
        raise SystemExit("No sheets found to clean.")
    expected = metas[0].raw_header
    mismatches = [
        f"{project_path(meta.source_path)}::{meta.sheet_name}"
        for meta in metas[1:]
        if meta.raw_header != expected
    ]
    if mismatches:
        details = "\n  ".join(mismatches)
        raise SystemExit(
            "The first three header rows are not identical across all selected sheets. "
            "No output was written.\nMismatched sheets:\n  "
            + details
        )


def parquet_schema(columns: list[str], date_col: str | None) -> pa.Schema:
    data: dict[str, pd.Series] = {}
    for column in columns:
        if column == date_col:
            continue
        if DATE_NAME_RE.search(str(column)):
            data[str(column)] = pd.Series(dtype="datetime64[ns]")
        elif STRING_NAME_RE.search(str(column)):
            data[str(column)] = pd.Series(dtype="string")
        else:
            data[str(column)] = pd.Series(dtype="float64")
    empty = pd.DataFrame(data)
    if date_col:
        empty.index = pd.DatetimeIndex([], name="date")
    return pa.Table.from_pandas(empty, preserve_index=bool(date_col)).schema


def coerce_to_schema(table: pa.Table, schema: pa.Schema) -> pa.Table:
    arrays = []
    for field in schema:
        if field.name in table.column_names:
            arrays.append(table[field.name].cast(field.type, safe=False))
        else:
            arrays.append(pa.nulls(table.num_rows, type=field.type))
    return pa.Table.from_arrays(arrays, schema=schema)


def clean_stream_batch(rows: list[tuple[Any, ...]], columns: list[str], date_col: str | None) -> pd.DataFrame:
    frame = pd.DataFrame.from_records(rows, columns=columns)
    frame = frame.dropna(how="all")
    if frame.empty:
        return frame
    for column in frame.columns:
        frame[column] = frame[column].map(clean_scalar)
    for column in frame.columns:
        name = str(column)
        if DATE_NAME_RE.search(name):
            parsed = pd.to_datetime(frame[column], errors="coerce")
            if parsed.notna().any():
                frame[column] = parsed
                continue
        if STRING_NAME_RE.search(name):
            frame[column] = frame[column].where(frame[column].notna(), pd.NA).astype("string").str.strip()
        else:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    if "Typrep" in frame.columns:
        frame = frame.loc[frame["Typrep"].astype("string").str.strip().eq("A")].copy()
    if date_col and date_col in frame.columns:
        frame[date_col] = pd.to_datetime(frame[date_col], errors="coerce")
    return frame


def iter_sheet_data_rows(path: Path, sheet_name: str) -> Any:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        sheet.reset_dimensions()
        for row in sheet.iter_rows(min_row=4, values_only=True):
            yield row
    finally:
        workbook.close()


def write_parquet_stream(
    metas: list[SheetMeta],
    output_path: Path,
    batch_size: int = 50_000,
) -> tuple[str | None, str, int, int]:
    validate_meta_headers(metas)
    columns = unique_column_names(metas[0].first_raw_values)
    date_cols = [meta.date_col for meta in metas if meta.date_col]
    date_col = date_cols[0] if date_cols and all(col == date_cols[0] for col in date_cols) else None
    schema = parquet_schema(columns, date_col)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    if temp_path.exists():
        temp_path.unlink()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    writer: pq.ParquetWriter | None = None
    date_values: list[Any] = []
    total_rows = 0
    total_cols = len(columns)

    try:
        writer = pq.ParquetWriter(temp_path, schema)
        for meta in metas:
            rows: list[tuple[Any, ...]] = []
            for row in iter_sheet_data_rows(meta.source_path, meta.sheet_name):
                if not any(normalize_cell(value) for value in row):
                    continue
                padded = tuple(row[: len(columns)]) + (None,) * max(0, len(columns) - len(row))
                rows.append(padded[: len(columns)])
                if len(rows) >= batch_size:
                    batch = clean_stream_batch(rows, columns, date_col)
                    rows = []
                    if batch.empty:
                        continue
                    if date_col and date_col in batch.columns:
                        date_values.extend(batch[date_col].dropna().tolist())
                        batch = batch.set_index(date_col)
                        batch.index.name = "date"
                    table = pa.Table.from_pandas(batch, preserve_index=bool(date_col))
                    writer.write_table(coerce_to_schema(table, schema))
                    total_rows += len(batch)
            if rows:
                batch = clean_stream_batch(rows, columns, date_col)
                if not batch.empty:
                    if date_col and date_col in batch.columns:
                        date_values.extend(batch[date_col].dropna().tolist())
                        batch = batch.set_index(date_col)
                        batch.index.name = "date"
                    table = pa.Table.from_pandas(batch, preserve_index=bool(date_col))
                    writer.write_table(coerce_to_schema(table, schema))
                    total_rows += len(batch)
        writer.close()
        writer = None

        table = pq.read_table(temp_path)
        if date_col and "date" in table.column_names:
            table = table.sort_by([("date", "ascending")])
        pq.write_table(table, output_path)
    finally:
        if writer is not None:
            writer.close()
        if temp_path.exists():
            temp_path.unlink()

    frequency = infer_frequency(pd.Series(date_values)) if date_col else "uncertain"
    for meta in metas:
        meta.frequency = frequency
    return date_col, frequency, total_rows, total_cols


def write_output(frame: pd.DataFrame, date_col: str | None, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    if suffix == ".parquet":
        out = frame.copy()
        if date_col and date_col in out.columns:
            out = out.set_index(date_col)
            out.index.name = "date"
        out.to_parquet(output_path)
        return
    if suffix in {".xlsx", ".xls"}:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name="Sheet1")
        return
    raise SystemExit(f"Unsupported output suffix: {output_path.suffix}. Use .xlsx, .xls, or .parquet.")


def _relative_to_a_data(path: Path) -> str | None:
    try:
        return str(path.resolve().relative_to(A_DATA_ROOT))
    except ValueError:
        return None


def update_inventory_a(
    inventory_path: Path,
    source_paths: list[Path],
    output_path: Path,
    frequency: str,
) -> int:
    """Update data_inventory_A.json: for each record whose file_path matches a source
    path, set clean_data and frequency. Returns the number of records updated."""
    if not inventory_path.exists():
        print(f"Warning: inventory file not found: {project_path(inventory_path)}")
        return 0

    with inventory_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    clean_rel = _relative_to_a_data(output_path)
    if clean_rel is None:
        clean_rel = str(output_path)

    source_rels = set()
    for sp in source_paths:
        rel = _relative_to_a_data(sp)
        if rel:
            source_rels.add(rel)

    updated = 0
    for sheet_data in data.get("sheets", {}).values():
        for record in sheet_data.get("records", []):
            fp = record.get("file_path", "")
            if fp in source_rels:
                record["clean_data"] = clean_rel
                record["frequency"] = frequency
                updated += 1

    inventory_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return updated


def sync_inventory_xlsx(inventory_path: Path) -> bool:
    """Sync the JSON inventory to a same-name .xlsx via convert_jsonl_to_xlsx.py.
    Returns True on success."""
    xlsx_path = inventory_path.with_suffix(".xlsx")
    if not CONVERT_SCRIPT.exists():
        print(f"Warning: convert script not found: {CONVERT_SCRIPT}")
        return False
    import importlib.util
    spec = importlib.util.spec_from_file_location("convert_jsonl_to_xlsx", CONVERT_SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    result = mod.convert_one_json(inventory_path, output_path=xlsx_path)
    if result["status"] == "success":
        return True
    print(f"Warning: failed to sync inventory xlsx: {result['error']}")
    return False


def source_stems(path: Path) -> set[str]:
    stem = path.stem
    return {stem, SPLIT_SUFFIX_RE.sub("", stem)}


def find_related_txt_files(source_paths: list[Path]) -> dict[Path, set[Path]]:
    search_dirs = {DATA_DIR}
    for path in source_paths:
        search_dirs.add(path.parent)
        # also search sibling "txt" directory (e.g. data_csmar/txt alongside data_csmar/xlsx)
        sibling_txt = path.parent.parent / "txt"
        if sibling_txt.is_dir():
            search_dirs.add(sibling_txt)

    txt_files = []
    for directory in search_dirs:
        if directory.exists():
            txt_files.extend(directory.glob("*.txt"))

    matches: dict[Path, set[Path]] = {path: set() for path in source_paths}
    for source_path in source_paths:
        stems = {stem for stem in source_stems(source_path) if stem}
        for txt_path in txt_files:
            txt_stem = txt_path.stem
            if any(stem in txt_stem or txt_stem in stem for stem in stems):
                matches[source_path].add(txt_path.resolve())
    return matches


def append_data_guide(txt_paths: list[Path], guide_path: Path) -> None:
    if not txt_paths:
        return
    guide_path.parent.mkdir(parents=True, exist_ok=True)
    existing = guide_path.read_text(encoding="utf-8") if guide_path.exists() else ""
    txt_paths = [path for path in txt_paths if f"## {path.name}" not in existing]
    if not txt_paths:
        return
    mode = "a" if guide_path.exists() else "w"
    with guide_path.open(mode, encoding="utf-8") as guide:
        if guide_path.stat().st_size > 0:
            guide.write("\n")
        for txt_path in txt_paths:
            text = txt_path.read_text(encoding="utf-8-sig")
            guide.write(f"## {txt_path.name}\n\n")
            guide.write(text.rstrip() + "\n\n")


def copy_txt_and_records(source_paths: list[Path], output_dir: Path) -> list[dict[str, Any]]:
    matches = find_related_txt_files(source_paths)
    copied: dict[Path, Path] = {}
    records: list[dict[str, Any]] = []
    guide_sources = []

    for source_path, txt_paths in matches.items():
        for txt_path in sorted(txt_paths, key=lambda p: p.name):
            destination = output_dir / txt_path.name
            output_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(txt_path, destination)
            if txt_path not in copied:
                copied[txt_path] = destination
                guide_sources.append(txt_path)
            records.append(
                {
                    "clean_data": txt_path.name,
                    "frequency": "uncertain",
                    "content": f"{source_path.name}的说明",
                    "data_source": project_path(txt_path),
                    "file_name": txt_path.name,
                    "sheet_name": "",
                    "first_raw_values": [],
                }
            )

    append_data_guide(guide_sources, output_dir / DATA_GUIDE_PATH.name)
    return records


def build_sheet_inventory_records(
    payloads: list[SheetPayload],
    output_name: str,
    content: str,
    frequency: str,
) -> list[dict[str, Any]]:
    records = []
    for payload in payloads:
        records.append(
            {
                "clean_data": output_name,
                "frequency": frequency if frequency != "uncertain" else payload.frequency,
                "content": content,
                "data_source": project_path(payload.source_path),
                "file_name": payload.source_path.name,
                "sheet_name": payload.sheet_name,
                "first_raw_values": payload.first_raw_values,
            }
        )
    return records


def build_meta_inventory_records(
    metas: list[SheetMeta],
    output_name: str,
    content: str,
    frequency: str,
) -> list[dict[str, Any]]:
    records = []
    for meta in metas:
        records.append(
            {
                "clean_data": output_name,
                "frequency": frequency if frequency != "uncertain" else meta.frequency,
                "content": content,
                "data_source": project_path(meta.source_path),
                "file_name": meta.source_path.name,
                "sheet_name": meta.sheet_name,
                "first_raw_values": meta.first_raw_values,
            }
        )
    return records


def main() -> None:
    args = parse_args()
    path_values = split_paths(args.files + args.files_option)
    if not path_values:
        path_values = prompt_for_paths()
    if not path_values:
        raise SystemExit("No input files provided.")

    source_paths = [resolve_input_path(value) for value in path_values]
    output_name = normalize_output_name(args.output or input("请输入目标输出文件名："))

    output_dir = args.output_dir.resolve()
    output_path = output_dir / output_name

    if output_path.suffix.lower() == ".parquet":
        metas = collect_sheet_metas(source_paths)
        date_col, frequency, row_count, col_count = write_parquet_stream(metas, output_path)
    else:
        payloads = load_payloads(source_paths)
        validate_headers(payloads)
        combined, date_col, frequency = combine_payloads(payloads)
        write_output(combined, date_col, output_path)
        row_count, col_count = combined.shape
    copy_txt_and_records(source_paths, output_dir)
    updated = update_inventory_a(INVENTORY_PATH, source_paths, output_path, frequency)
    xlsx_synced = sync_inventory_xlsx(INVENTORY_PATH)

    print(f"Wrote {project_path(output_path)} ({row_count} rows x {col_count} columns)")
    print(f"Updated {updated} record(s) in {project_path(INVENTORY_PATH)}")
    if xlsx_synced:
        print(f"Synced {project_path(INVENTORY_PATH.with_suffix('.xlsx'))}")
    if (output_dir / DATA_GUIDE_PATH.name).exists():
        print(f"Wrote/updated {project_path(output_dir / DATA_GUIDE_PATH.name)}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("\nInterrupted.")
