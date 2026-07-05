"""为基金面板追加状态匹配未来 6 期收益率。

默认读取 ``panel_base.parquet`` 一次完成计算，再按 ``ifind_code`` 和自然月
主键对齐写入基准面板与热力图面板。对每个 t 月和每个 regime，从 t+1 开始
寻找最大前看范围内最近 6 个属于该 regime 的月份。每个目标月 tau 的单月
收益要求 tau-1 与 tau 两端记录都存在、月份连续、NAV 有效，且两端的
``is_sample`` 与 ``is_size_eligible_t`` 都为 True；任一目标月无效时整段收益
记为 NaN，不会改用更远月份替补。

原版 ``is_insample_future_ret_6m`` 只按当前月加 6 个月后的理论终点是否不晚于
``PANEL_INSAMPLE_END_DATE`` 判断，与收益是否缺失无关。状态匹配版采用等价口径：
只要实际选出的第 6 个状态月不晚于截止月就记 1，否则记 0；基金层面 NAV 或
资格条件导致 Y 缺失时，标记仍由第 6 个状态月独立决定。若找不到第 6 个状态
月，则条件不成立，标记为 0。

写盘采用临时文件、完整回读校验和 ``os.replace`` 原子替换。脚本可重复运行，
每次会先从目标表中移除本脚本拥有的旧列，再追加本次结果。

运行方式：

    .venv/bin/python A_data/scripts/3_panel_base_mkt_condition_future_returns.py
    .venv/bin/python A_data/scripts/3_panel_base_mkt_condition_future_returns.py --max-forward 48
"""

from __future__ import annotations

import argparse
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.compute as pc
import pyarrow.parquet as pq
from openpyxl.styles import Font, PatternFill

import Config


DEFAULT_SOURCE_PANEL_PATH = Config.PANEL_OUTPUT_PATH
DEFAULT_TARGET_PATHS = (
    Config.PANEL_OUTPUT_PATH,
    Config.PANEL_HEATMAP_OUTPUT_PATH,
)
DEFAULT_CHECK_OUTPUT_PATH = (
    Config.A_DATA_ROOT / "output" / "mkt_condition_future_returns_check.xlsx"
)
DEFAULT_MAX_FORWARD_MONTHS = 36
DEFAULT_PREVIEW_ROWS = 1000
FUTURE_STATE_MONTH_COUNT = 6
MISSING_ORDINAL = -1
BRUTE_FORCE_SAMPLE_SIZE = 100
BRUTE_FORCE_RANDOM_SEED = 20260702
MIN_CROSS_SECTION_N = 50
PLAIN_FUTURE_RETURN_COLUMN = "future_ret_6m"

STATE_REGIME_DEFINITIONS = (
    ("mkt_state_hs300", (("hs300up", 1.0), ("hs300down", -1.0))),
    ("mkt_state_style", (("growth", 1.0), ("value", -1.0))),
    ("mkt_state_size", (("large", 1.0), ("small", -1.0))),
    ("mkt_state_indvol", (("highvol", 1.0), ("lowvol", -1.0))),
)
STATE_COLUMNS = tuple(state for state, _ in STATE_REGIME_DEFINITIONS)


@dataclass
class PreparedPanelUpdate:
    """保存已写入并校验、但尚未替换正式文件的临时产物。"""

    panel_path: Path
    temporary_panel_path: Path
    preview_path: Path | None
    temporary_preview_path: Path | None
    row_count: int
    column_count: int
    replaced_existing_columns: int


def iter_regimes() -> Iterable[tuple[str, str, float]]:
    """依次给出 (状态列, regime 名, 状态取值)。"""
    for state_column, regimes in STATE_REGIME_DEFINITIONS:
        for regime, state_value in regimes:
            yield state_column, regime, state_value


def get_return_column(regime: str) -> str:
    """返回一个 regime 的状态匹配未来收益列名。"""
    return f"future_ret_6m_{regime}"


def get_insample_column(regime: str) -> str:
    """返回一个 regime 的样本内标记列名。"""
    return f"is_insample_future_ret_6m_{regime}"


def get_owned_output_columns() -> list[str]:
    """列出本脚本拥有的 16 列，供重复运行时删除旧版本。"""
    columns: list[str] = []
    for _, regime, _ in iter_regimes():
        columns.extend([get_return_column(regime), get_insample_column(regime)])
    return columns


def parse_args() -> argparse.Namespace:
    """读取命令行参数。"""
    parser = argparse.ArgumentParser(
        description="为两个基金面板追加状态匹配未来 6 期收益率。"
    )
    parser.add_argument(
        "--source-panel",
        type=Path,
        default=DEFAULT_SOURCE_PANEL_PATH,
        help="计算结果使用的 panel_base parquet 路径。",
    )
    parser.add_argument(
        "--targets",
        type=Path,
        nargs="+",
        default=list(DEFAULT_TARGET_PATHS),
        help="需要追加结果列的 parquet 路径，默认写入两个正式面板。",
    )
    parser.add_argument(
        "--max-forward",
        type=int,
        default=DEFAULT_MAX_FORWARD_MONTHS,
        help="寻找 6 个状态月时允许的最大前看月数，默认 36。",
    )
    parser.add_argument(
        "--check-output",
        type=Path,
        default=DEFAULT_CHECK_OUTPUT_PATH,
        help="校验工作簿输出路径。",
    )
    parser.add_argument(
        "--preview-rows",
        type=int,
        default=DEFAULT_PREVIEW_ROWS,
        help="parquet 对应预览文件保留的前部行数。",
    )
    parser.add_argument(
        "--no-preview",
        action="store_true",
        help="只写 parquet，不刷新 Excel 预览。",
    )
    return parser.parse_args()


def read_source_panel(source_panel_path: Path) -> pd.DataFrame:
    """读取计算所需的最小字段集合，并统一月份与数值类型。"""
    if not source_panel_path.exists():
        raise FileNotFoundError(f"面板文件不存在：{source_panel_path}")

    required_columns = [
        Config.COLUMN_IFIND_CODE,
        Config.COLUMN_MONTH_DATE,
        Config.COLUMN_NAV,
        Config.COLUMN_IS_SAMPLE,
        Config.COLUMN_IS_SIZE_ELIGIBLE,
        PLAIN_FUTURE_RETURN_COLUMN,
        *STATE_COLUMNS,
    ]
    schema_names = set(pq.read_schema(source_panel_path).names)
    missing = [column for column in required_columns if column not in schema_names]
    if missing:
        raise ValueError(
            f"{source_panel_path} 缺少字段：{missing}；"
            "请先运行 3_generate_panel_base.py 和 3_panel_base_mkt_condition.py。"
        )

    data = pd.read_parquet(source_panel_path, columns=required_columns)
    data[Config.COLUMN_MONTH_DATE] = pd.to_datetime(
        data[Config.COLUMN_MONTH_DATE], errors="coerce"
    )
    if data[Config.COLUMN_MONTH_DATE].isna().any():
        raise ValueError(f"{source_panel_path} 存在无法识别的 month_date。")
    if data.duplicated(
        [Config.COLUMN_IFIND_CODE, Config.COLUMN_MONTH_DATE]
    ).any():
        raise AssertionError(f"{source_panel_path} 存在重复基金-月份键。")

    data[Config.COLUMN_NAV] = pd.to_numeric(
        data[Config.COLUMN_NAV], errors="coerce"
    )
    data[Config.COLUMN_IS_SAMPLE] = (
        data[Config.COLUMN_IS_SAMPLE].fillna(False).astype(bool)
    )
    data[Config.COLUMN_IS_SIZE_ELIGIBLE] = (
        data[Config.COLUMN_IS_SIZE_ELIGIBLE].fillna(False).astype(bool)
    )
    for state_column in STATE_COLUMNS:
        data[state_column] = pd.to_numeric(data[state_column], errors="coerce")
        if not data[state_column].dropna().isin([-1.0, 1.0]).all():
            raise AssertionError(f"{state_column} 出现 1/-1 之外的非缺失值。")
    return data


def extract_month_state_table(data: pd.DataFrame) -> pd.DataFrame:
    """提取每个自然月唯一的市场状态，并校验同月基金状态完全一致。"""
    month_periods = data[Config.COLUMN_MONTH_DATE].dt.to_period("M")
    for state_column in STATE_COLUMNS:
        unique_counts = data.groupby(month_periods, sort=False)[state_column].nunique(
            dropna=False
        )
        if (unique_counts != 1).any():
            bad_months = unique_counts[unique_counts != 1].index.astype(str).tolist()
            raise AssertionError(
                f"{state_column} 在同一月份并非所有基金取值一致：{bad_months[:5]}"
            )

    state_table = data.groupby(month_periods, sort=True)[list(STATE_COLUMNS)].first()
    state_table.index.name = "month_period"
    if not state_table.index.is_monotonic_increasing:
        raise AssertionError("月份状态表未按自然月递增排序。")
    return state_table


def build_target_months(
    state_table: pd.DataFrame,
    max_forward: int,
) -> tuple[dict[str, np.ndarray], pd.DataFrame]:
    """为每个 t 月和 regime 选择最近 6 个未来状态月。"""
    month_ordinals = state_table.index.asi8.astype("int64", copy=False)
    target_by_regime: dict[str, np.ndarray] = {}
    detail_rows: list[dict[str, object]] = []

    for state_column, regime, state_value in iter_regimes():
        state_values = state_table[state_column].to_numpy(dtype="float64")
        regime_months = month_ordinals[state_values == state_value]
        target_matrix = np.full(
            (len(month_ordinals), FUTURE_STATE_MONTH_COUNT),
            MISSING_ORDINAL,
            dtype="int64",
        )

        # 月份数量很少，按 t 月逐个切片比构造庞大的笛卡尔积更清楚也更省内存。
        for month_position, current_ordinal in enumerate(month_ordinals):
            first_future = int(np.searchsorted(regime_months, current_ordinal + 1))
            candidates = regime_months[
                first_future : first_future + FUTURE_STATE_MONTH_COUNT
            ]
            candidates = candidates[candidates <= current_ordinal + max_forward]
            target_matrix[month_position, : len(candidates)] = candidates

            sixth_ordinal = target_matrix[month_position, -1]
            has_sixth = sixth_ordinal != MISSING_ORDINAL
            detail_rows.append(
                {
                    "month_date": state_table.index[month_position].to_timestamp("M"),
                    "regime": regime,
                    "sixth_future_state_month": (
                        pd.Period(ordinal=int(sixth_ordinal), freq="M").to_timestamp(
                            "M"
                        )
                        if has_sixth
                        else pd.NaT
                    ),
                    "calendar_span_months": (
                        int(sixth_ordinal - current_ordinal) if has_sixth else np.nan
                    ),
                    "has_six_future_state_months": int(has_sixth),
                }
            )
        target_by_regime[regime] = target_matrix

    return target_by_regime, pd.DataFrame(detail_rows)


def build_panel_lookup(data: pd.DataFrame) -> pd.DataFrame:
    """建立基金代码 x 月份 ordinal 查找表，供向量化与暴力复算共用。"""
    month_ordinals = data[Config.COLUMN_MONTH_DATE].dt.to_period("M").array.asi8
    lookup = data[
        [
            Config.COLUMN_IFIND_CODE,
            Config.COLUMN_NAV,
            Config.COLUMN_IS_SAMPLE,
            Config.COLUMN_IS_SIZE_ELIGIBLE,
        ]
    ].copy()
    lookup["month_ordinal"] = month_ordinals
    lookup = lookup.set_index([Config.COLUMN_IFIND_CODE, "month_ordinal"])
    if lookup.index.duplicated().any():
        raise AssertionError("基金-月份 ordinal 查找表存在重复键。")
    return lookup


def reindex_lookup_column(
    lookup: pd.DataFrame,
    fund_codes: np.ndarray,
    month_ordinals: np.ndarray,
    column: str,
) -> np.ndarray:
    """按基金和目标月批量查找一列，缺失键自然返回 NaN。"""
    query_index = pd.MultiIndex.from_arrays(
        [fund_codes, month_ordinals],
        names=[Config.COLUMN_IFIND_CODE, "month_ordinal"],
    )
    return lookup[column].reindex(query_index).to_numpy()


def calculate_regime_returns(
    data: pd.DataFrame,
    state_table: pd.DataFrame,
    target_by_regime: dict[str, np.ndarray],
) -> tuple[pd.DataFrame, dict[str, np.ndarray]]:
    """按已选目标月计算 8 组状态匹配收益率和样本内标记。"""
    month_ordinals = data[Config.COLUMN_MONTH_DATE].dt.to_period("M").array.asi8
    # 直接用 PeriodIndex 的整数 ordinal 对齐，避免 pandas 版本间构造参数差异。
    month_positions = pd.Index(state_table.index.asi8).get_indexer(month_ordinals)
    if (month_positions < 0).any():
        raise AssertionError("面板中存在月份状态表未覆盖的月份。")

    fund_codes = data[Config.COLUMN_IFIND_CODE].to_numpy()
    lookup = build_panel_lookup(data)
    insample_end_ordinal = pd.Period(
        pd.Timestamp(Config.PANEL_INSAMPLE_END_DATE), freq="M"
    ).ordinal
    output: dict[str, pd.Series] = {}
    row_targets: dict[str, np.ndarray] = {}

    for _, regime, _ in iter_regimes():
        targets = target_by_regime[regime][month_positions]
        row_targets[regime] = targets
        gross_returns = np.ones(len(data), dtype="float64")
        all_legs_valid = targets[:, -1] != MISSING_ORDINAL

        for target_position in range(FUTURE_STATE_MONTH_COUNT):
            target_ordinals = targets[:, target_position]
            previous_ordinals = np.where(
                target_ordinals == MISSING_ORDINAL,
                MISSING_ORDINAL,
                target_ordinals - 1,
            )

            target_nav = reindex_lookup_column(
                lookup, fund_codes, target_ordinals, Config.COLUMN_NAV
            ).astype("float64")
            previous_nav = reindex_lookup_column(
                lookup, fund_codes, previous_ordinals, Config.COLUMN_NAV
            ).astype("float64")
            target_sample = reindex_lookup_column(
                lookup, fund_codes, target_ordinals, Config.COLUMN_IS_SAMPLE
            )
            previous_sample = reindex_lookup_column(
                lookup, fund_codes, previous_ordinals, Config.COLUMN_IS_SAMPLE
            )
            target_size = reindex_lookup_column(
                lookup,
                fund_codes,
                target_ordinals,
                Config.COLUMN_IS_SIZE_ELIGIBLE,
            )
            previous_size = reindex_lookup_column(
                lookup,
                fund_codes,
                previous_ordinals,
                Config.COLUMN_IS_SIZE_ELIGIBLE,
            )

            # 查找必须精确落在 tau-1 和 tau，因此键存在本身就证明月份连续。
            leg_valid = (
                (target_ordinals != MISSING_ORDINAL)
                & np.isfinite(target_nav)
                & np.isfinite(previous_nav)
                & (target_nav > 0)
                & (previous_nav > 0)
                & (target_sample == True)
                & (previous_sample == True)
                & (target_size == True)
                & (previous_size == True)
            )
            all_legs_valid &= leg_valid
            gross_returns *= np.where(leg_valid, target_nav / previous_nav, 1.0)

        values = np.full(len(data), np.nan, dtype="float64")
        values[all_legs_valid] = gross_returns[all_legs_valid] - 1.0
        sixth_target = targets[:, -1]
        insample_values = (
            (sixth_target != MISSING_ORDINAL)
            & (sixth_target <= insample_end_ordinal)
        ).astype("int64")

        output[get_return_column(regime)] = pd.Series(values, index=data.index)
        output[get_insample_column(regime)] = pd.Series(
            insample_values, index=data.index, dtype="int64"
        )

    return pd.DataFrame(output, index=data.index), row_targets


def validate_identity(
    data: pd.DataFrame,
    state_table: pd.DataFrame,
    output_columns: pd.DataFrame,
) -> dict[str, int]:
    """全量检查连续 6 个日历月同市态时与原未来 6 月收益相等。"""
    row_ordinals = data[Config.COLUMN_MONTH_DATE].dt.to_period("M").array.asi8
    state_ordinals = state_table.index.asi8
    plain_return = pd.to_numeric(
        data[PLAIN_FUTURE_RETURN_COLUMN], errors="coerce"
    ).to_numpy(dtype="float64")
    coverage: dict[str, int] = {}

    for state_column, regime, state_value in iter_regimes():
        state_lookup = dict(
            zip(state_ordinals.tolist(), state_table[state_column].tolist())
        )
        identity_months = {
            int(current_ordinal)
            for current_ordinal in state_ordinals
            if all(
                state_lookup.get(int(current_ordinal + offset), np.nan)
                == state_value
                for offset in range(1, FUTURE_STATE_MONTH_COUNT + 1)
            )
        }
        month_mask = np.isin(row_ordinals, np.fromiter(identity_months, dtype="int64"))
        regime_return = output_columns[get_return_column(regime)].to_numpy(
            dtype="float64"
        )
        compare_mask = month_mask & np.isfinite(plain_return) & np.isfinite(
            regime_return
        )
        checked_rows = int(compare_mask.sum())
        if checked_rows and not np.isclose(
            regime_return[compare_mask],
            plain_return[compare_mask],
            rtol=1e-10,
            atol=1e-12,
        ).all():
            differences = np.abs(
                regime_return[compare_mask] - plain_return[compare_mask]
            )
            raise AssertionError(
                f"{regime} 恒等式检查失败，最大绝对误差为 {differences.max()}。"
            )
        coverage[regime] = checked_rows
    return coverage


def validate_nan_and_flags(
    output_columns: pd.DataFrame,
    row_targets: dict[str, np.ndarray],
) -> None:
    """检查目标月不足时的整月缺失，以及标记的类型与独立性规则。"""
    insample_end_ordinal = pd.Period(
        pd.Timestamp(Config.PANEL_INSAMPLE_END_DATE), freq="M"
    ).ordinal
    for _, regime, _ in iter_regimes():
        values = output_columns[get_return_column(regime)]
        flags = output_columns[get_insample_column(regime)]
        sixth_target = row_targets[regime][:, -1]
        missing_targets = sixth_target == MISSING_ORDINAL
        if values[missing_targets].notna().any():
            raise AssertionError(f"{regime} 目标月不足时出现非缺失收益。")
        if flags.dtype != np.dtype("int64"):
            raise AssertionError(f"{regime} 的样本内标记不是 int64。")
        if not flags.isin([0, 1]).all():
            raise AssertionError(f"{regime} 的样本内标记出现 0/1 之外的值。")
        expected_flags = (
            (~missing_targets) & (sixth_target <= insample_end_ordinal)
        ).astype("int64")
        if not np.array_equal(flags.to_numpy(), expected_flags):
            raise AssertionError(f"{regime} 的样本内标记与第 6 个状态月不一致。")


def brute_force_check(
    data: pd.DataFrame,
    state_table: pd.DataFrame,
    output_columns: pd.DataFrame,
    max_forward: int,
) -> int:
    """固定随机种子抽 100 个非缺失结果，用纯 Python 循环独立复算。"""
    record_lookup = {
        (row.ifind_code, int(row.month_ordinal)): (
            float(row.nav) if pd.notna(row.nav) else np.nan,
            bool(row.is_sample),
            bool(row.is_size_eligible),
        )
        for row in pd.DataFrame(
            {
                Config.COLUMN_IFIND_CODE: data[Config.COLUMN_IFIND_CODE],
                "month_ordinal": data[Config.COLUMN_MONTH_DATE]
                .dt.to_period("M")
                .array.asi8,
                "nav": data[Config.COLUMN_NAV],
                "is_sample": data[Config.COLUMN_IS_SAMPLE],
                "is_size_eligible": data[Config.COLUMN_IS_SIZE_ELIGIBLE],
            }
        ).itertuples(index=False)
    }
    state_months: dict[str, list[int]] = {}
    for state_column, regime, state_value in iter_regimes():
        state_months[regime] = [
            int(ordinal)
            for ordinal, value in zip(
                state_table.index.asi8, state_table[state_column].to_numpy()
            )
            if value == state_value
        ]

    candidate_regimes: list[str] = []
    candidate_rows: list[int] = []
    for _, regime, _ in iter_regimes():
        positions = np.flatnonzero(
            output_columns[get_return_column(regime)].notna().to_numpy()
        )
        candidate_regimes.extend([regime] * len(positions))
        candidate_rows.extend(positions.tolist())
    if len(candidate_rows) < BRUTE_FORCE_SAMPLE_SIZE:
        raise AssertionError("非缺失结果不足 100 个，无法完成暴力复算。")

    rng = np.random.default_rng(BRUTE_FORCE_RANDOM_SEED)
    sampled_positions = rng.choice(
        len(candidate_rows), size=BRUTE_FORCE_SAMPLE_SIZE, replace=False
    )
    for sample_position in sampled_positions:
        regime = candidate_regimes[int(sample_position)]
        row_position = candidate_rows[int(sample_position)]
        row = data.iloc[row_position]
        current_ordinal = int(
            pd.Period(row[Config.COLUMN_MONTH_DATE], freq="M").ordinal
        )
        selected_months = [
            ordinal
            for ordinal in state_months[regime]
            if current_ordinal < ordinal <= current_ordinal + max_forward
        ][:FUTURE_STATE_MONTH_COUNT]
        if len(selected_months) != FUTURE_STATE_MONTH_COUNT:
            raise AssertionError(f"{regime} 暴力复算抽到目标月不足的非缺失行。")

        gross_return = 1.0
        fund_code = row[Config.COLUMN_IFIND_CODE]
        for target_ordinal in selected_months:
            previous_record = record_lookup.get((fund_code, target_ordinal - 1))
            target_record = record_lookup.get((fund_code, target_ordinal))
            if previous_record is None or target_record is None:
                raise AssertionError(f"{regime} 暴力复算抽到缺少相邻月份的非缺失行。")
            previous_nav, previous_sample, previous_size = previous_record
            target_nav, target_sample, target_size = target_record
            if not (
                np.isfinite(previous_nav)
                and np.isfinite(target_nav)
                and previous_nav > 0
                and target_nav > 0
                and previous_sample
                and target_sample
                and previous_size
                and target_size
            ):
                raise AssertionError(f"{regime} 暴力复算抽到资格条件不全的非缺失行。")
            gross_return *= target_nav / previous_nav

        expected = gross_return - 1.0
        actual = output_columns.iloc[row_position][get_return_column(regime)]
        if not np.isclose(actual, expected, rtol=1e-10, atol=1e-12):
            raise AssertionError(
                f"{regime} 行 {row_position} 暴力复算不一致：{actual} vs {expected}。"
            )
    return BRUTE_FORCE_SAMPLE_SIZE


def build_summary(
    data: pd.DataFrame,
    output_columns: pd.DataFrame,
    month_detail: pd.DataFrame,
    identity_coverage: dict[str, int],
) -> pd.DataFrame:
    """汇总每个 regime 的覆盖、有效月份、窗口跨度与恒等式覆盖。"""
    month_periods = data[Config.COLUMN_MONTH_DATE].dt.to_period("M")
    rows: list[dict[str, object]] = []
    for _, regime, _ in iter_regimes():
        values = output_columns[get_return_column(regime)]
        flags = output_columns[get_insample_column(regime)] == 1
        insample_counts = values[flags].notna().groupby(month_periods[flags]).sum()
        detail = month_detail[month_detail["regime"] == regime]
        spans = pd.to_numeric(detail["calendar_span_months"], errors="coerce")
        rows.append(
            {
                "regime": regime,
                "non_null_rows": int(values.notna().sum()),
                "non_null_rows_insample": int((values.notna() & flags).sum()),
                "effective_insample_months_ge50": int(
                    (insample_counts >= MIN_CROSS_SECTION_N).sum()
                ),
                "mean_calendar_span_months": float(spans.mean()),
                "max_calendar_span_months": (
                    int(spans.max()) if spans.notna().any() else np.nan
                ),
                "months_missing_six_state_months": int(
                    (detail["has_six_future_state_months"] == 0).sum()
                ),
                "identity_check_rows": int(identity_coverage[regime]),
            }
        )
    return pd.DataFrame(rows)


def make_temporary_path(target_path: Path) -> Path:
    """在目标目录建立同后缀临时文件，确保之后可以原子替换。"""
    target_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{target_path.stem}_",
        suffix=target_path.suffix,
        dir=target_path.parent,
    )
    os.close(descriptor)
    return Path(temporary_name)


def get_default_preview_path(panel_path: Path) -> Path:
    """按项目约定推导 parquet 的 Excel 预览路径。"""
    return panel_path.with_name(panel_path.stem + "_preview.xlsx")


def validate_original_columns_after_write(
    original_table: pa.Table,
    written_path: Path,
    original_column_names: list[str],
) -> None:
    """分批回读临时 parquet，逐列逐行确认所有原有字段完全不变。"""
    written_file = pq.ParquetFile(written_path)
    row_offset = 0
    original_only = original_table.select(original_column_names)
    for row_group_index in range(written_file.num_row_groups):
        written_group = written_file.read_row_group(
            row_group_index, columns=original_column_names
        )
        expected_group = original_only.slice(row_offset, written_group.num_rows)
        for column_name in original_column_names:
            actual = written_group[column_name]
            expected = expected_group[column_name]
            if actual.type != expected.type:
                raise AssertionError(
                    f"{written_path} 的原有列 {column_name} 回读类型发生变化。"
                )

            # Arrow 的 Table.equals 会把 NaN 与 NaN 判为不相等。这里显式把
            # 双方同时为 null 或同时为 NaN 视作相等，再检查其余实际值。
            both_null = pc.and_(pc.is_null(actual), pc.is_null(expected))
            equal_values = pc.fill_null(pc.equal(actual, expected), False)
            if pa.types.is_floating(actual.type):
                both_nan = pc.fill_null(
                    pc.and_(pc.is_nan(actual), pc.is_nan(expected)), False
                )
                equal_values = pc.or_(equal_values, both_nan)
            all_equal = pc.all(pc.or_(both_null, equal_values)).as_py()
            if not all_equal:
                raise AssertionError(
                    f"{written_path} 回读后原有列 {column_name} 在第 "
                    f"{row_group_index} 个行组发生变化。"
                )
        row_offset += written_group.num_rows
    if row_offset != original_table.num_rows:
        raise AssertionError(f"{written_path} 回读累计行数与原表不一致。")


def prepare_panel_update(
    panel_path: Path,
    source_keys: pd.MultiIndex,
    output_columns: pd.DataFrame,
    preview_rows: int,
    write_preview: bool,
) -> PreparedPanelUpdate:
    """生成、回读并校验单个目标面板的临时文件，但暂不替换正式文件。"""
    if not panel_path.exists():
        raise FileNotFoundError(f"目标面板不存在：{panel_path}")
    key_table = pq.read_table(
        panel_path,
        columns=[Config.COLUMN_IFIND_CODE, Config.COLUMN_MONTH_DATE],
    ).to_pandas()
    target_keys = pd.MultiIndex.from_arrays(
        [
            key_table[Config.COLUMN_IFIND_CODE].to_numpy(),
            pd.to_datetime(key_table[Config.COLUMN_MONTH_DATE])
            .dt.to_period("M")
            .array.asi8,
        ],
        names=source_keys.names,
    )
    if target_keys.duplicated().any():
        raise AssertionError(f"{panel_path} 存在重复基金-月份键。")
    if len(target_keys) != len(source_keys):
        raise AssertionError(f"{panel_path} 与来源面板行数不一致。")
    if len(target_keys.difference(source_keys)) or len(
        source_keys.difference(target_keys)
    ):
        raise AssertionError(f"{panel_path} 与来源面板的基金-月份键集合不一致。")

    aligned = output_columns.set_axis(source_keys, axis=0).reindex(target_keys)
    owned_columns = set(get_owned_output_columns())
    table = pq.read_table(panel_path)
    original_column_names = [
        name for name in table.column_names if name not in owned_columns
    ]
    replaced_columns = [
        name for name in table.column_names if name in owned_columns
    ]
    updated_table = table.select(original_column_names)
    for column in get_owned_output_columns():
        if column.startswith("is_insample_"):
            arrow_values = pa.array(
                aligned[column].to_numpy(dtype="int64"), type=pa.int64()
            )
        else:
            arrow_values = pa.array(
                aligned[column].to_numpy(dtype="float64"), type=pa.float64()
            )
        updated_table = updated_table.append_column(column, arrow_values)

    temporary_panel = make_temporary_path(panel_path)
    preview_path = get_default_preview_path(panel_path) if write_preview else None
    temporary_preview = (
        make_temporary_path(preview_path) if preview_path is not None else None
    )
    try:
        pq.write_table(updated_table, temporary_panel, row_group_size=50_000)
        written_file = pq.ParquetFile(temporary_panel)
        if written_file.metadata.num_rows != updated_table.num_rows:
            raise AssertionError("临时 parquet 回读行数不一致。")
        if written_file.metadata.num_columns != updated_table.num_columns:
            raise AssertionError("临时 parquet 回读列数不一致。")
        if written_file.schema_arrow.names != updated_table.column_names:
            raise AssertionError("临时 parquet 回读列名或顺序不一致。")

        validate_original_columns_after_write(
            table, temporary_panel, original_column_names
        )
        written_keys = pq.read_table(
            temporary_panel,
            columns=[Config.COLUMN_IFIND_CODE, Config.COLUMN_MONTH_DATE],
        ).to_pandas()
        written_key_index = pd.MultiIndex.from_arrays(
            [
                written_keys[Config.COLUMN_IFIND_CODE].to_numpy(),
                pd.to_datetime(written_keys[Config.COLUMN_MONTH_DATE])
                .dt.to_period("M")
                .array.asi8,
            ],
            names=target_keys.names,
        )
        if not written_key_index.equals(target_keys):
            raise AssertionError("临时 parquet 回读后主键或行顺序发生变化。")

        # 再回读新增列，避免只验证元数据而漏掉类型或数据写入问题。
        written_outputs = pq.read_table(
            temporary_panel, columns=get_owned_output_columns()
        ).to_pandas()
        for column in get_owned_output_columns():
            expected = aligned[column].to_numpy()
            actual = written_outputs[column].to_numpy()
            if column.startswith("is_insample_"):
                if actual.dtype != np.dtype("int64") or not np.array_equal(
                    actual, expected
                ):
                    raise AssertionError(f"临时 parquet 的 {column} 回读不一致。")
            elif not np.allclose(actual, expected, equal_nan=True):
                raise AssertionError(f"临时 parquet 的 {column} 回读不一致。")

        if temporary_preview is not None:
            preview = updated_table.slice(0, preview_rows).to_pandas()
            preview.to_excel(temporary_preview, index=False)
            preview_check = pd.read_excel(temporary_preview, nrows=5)
            if list(preview_check.columns) != list(preview.columns):
                raise AssertionError("临时 Excel 预览回读字段顺序不一致。")

        return PreparedPanelUpdate(
            panel_path=panel_path,
            temporary_panel_path=temporary_panel,
            preview_path=preview_path,
            temporary_preview_path=temporary_preview,
            row_count=updated_table.num_rows,
            column_count=updated_table.num_columns,
            replaced_existing_columns=len(replaced_columns),
        )
    except Exception:
        temporary_panel.unlink(missing_ok=True)
        if temporary_preview is not None:
            temporary_preview.unlink(missing_ok=True)
        raise


def prepare_check_workbook(
    summary: pd.DataFrame,
    month_detail: pd.DataFrame,
    check_output_path: Path,
) -> Path:
    """写出并回读校验工作簿临时文件，返回待原子替换的路径。"""
    temporary_check = make_temporary_path(check_output_path)
    try:
        with pd.ExcelWriter(temporary_check, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="summary", index=False)
            month_detail.to_excel(writer, sheet_name="month_detail", index=False)
            for sheet_name, frame in (
                ("summary", summary),
                ("month_detail", month_detail),
            ):
                worksheet = writer.book[sheet_name]
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
                for row_cells in worksheet.iter_rows(
                    min_row=2, max_row=len(frame) + 1
                ):
                    for cell in row_cells:
                        cell.fill = PatternFill(
                            fill_type="solid", fgColor="FFFFFF"
                        )
                for column_cells in worksheet.columns:
                    values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                    width = min(max(max(map(len, values)) + 2, 12), 32)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = width
                if sheet_name == "summary":
                    for row in range(2, len(frame) + 2):
                        worksheet.cell(row, 5).number_format = "0.00"
                else:
                    for row in range(2, len(frame) + 2):
                        worksheet.cell(row, 1).number_format = "yyyy-mm-dd"
                        worksheet.cell(row, 3).number_format = "yyyy-mm-dd"

        read_summary = pd.read_excel(temporary_check, sheet_name="summary")
        read_detail = pd.read_excel(temporary_check, sheet_name="month_detail")
        if list(read_summary.columns) != list(summary.columns) or len(
            read_summary
        ) != len(summary):
            raise AssertionError("校验表 summary 回读结构不一致。")
        if list(read_detail.columns) != list(month_detail.columns) or len(
            read_detail
        ) != len(month_detail):
            raise AssertionError("校验表 month_detail 回读结构不一致。")
        return temporary_check
    except Exception:
        temporary_check.unlink(missing_ok=True)
        raise


def commit_prepared_outputs(
    panel_updates: list[PreparedPanelUpdate],
    temporary_check: Path,
    check_output_path: Path,
) -> None:
    """在所有临时产物均通过校验后，以原子替换提交正式文件。"""
    try:
        for update in panel_updates:
            os.replace(update.temporary_panel_path, update.panel_path)
            if (
                update.temporary_preview_path is not None
                and update.preview_path is not None
            ):
                os.replace(update.temporary_preview_path, update.preview_path)
        os.replace(temporary_check, check_output_path)
    finally:
        for update in panel_updates:
            update.temporary_panel_path.unlink(missing_ok=True)
            if update.temporary_preview_path is not None:
                update.temporary_preview_path.unlink(missing_ok=True)
        temporary_check.unlink(missing_ok=True)


def generate_market_condition_future_returns(
    source_panel_path: Path = DEFAULT_SOURCE_PANEL_PATH,
    target_paths: Iterable[Path] = DEFAULT_TARGET_PATHS,
    max_forward: int = DEFAULT_MAX_FORWARD_MONTHS,
    check_output_path: Path = DEFAULT_CHECK_OUTPUT_PATH,
    preview_rows: int = DEFAULT_PREVIEW_ROWS,
    write_preview: bool = True,
) -> dict[str, object]:
    """完成计算、写盘前验证、双面板临时写入与原子替换。"""
    if max_forward < FUTURE_STATE_MONTH_COUNT:
        raise ValueError(
            f"max_forward 不得小于 {FUTURE_STATE_MONTH_COUNT}。"
        )
    if preview_rows <= 0:
        raise ValueError("preview_rows 必须是正整数。")

    data = read_source_panel(source_panel_path)
    state_table = extract_month_state_table(data)
    target_by_regime, month_detail = build_target_months(
        state_table, max_forward
    )
    output_columns, row_targets = calculate_regime_returns(
        data, state_table, target_by_regime
    )
    if list(output_columns.columns) != get_owned_output_columns():
        raise AssertionError("新增列顺序与本脚本登记的列清单不一致。")

    # 所有会影响数据正确性的检查必须在准备任何正式写盘之前完成。
    identity_coverage = validate_identity(data, state_table, output_columns)
    validate_nan_and_flags(output_columns, row_targets)
    brute_force_rows = brute_force_check(
        data, state_table, output_columns, max_forward
    )
    summary = build_summary(
        data, output_columns, month_detail, identity_coverage
    )

    source_keys = pd.MultiIndex.from_arrays(
        [
            data[Config.COLUMN_IFIND_CODE].to_numpy(),
            data[Config.COLUMN_MONTH_DATE].dt.to_period("M").array.asi8,
        ],
        names=[Config.COLUMN_IFIND_CODE, "month_ordinal"],
    )
    if source_keys.duplicated().any():
        raise AssertionError("来源面板主键重复。")

    panel_updates: list[PreparedPanelUpdate] = []
    temporary_check: Path | None = None
    try:
        for target_path in target_paths:
            panel_updates.append(
                prepare_panel_update(
                    Path(target_path),
                    source_keys,
                    output_columns,
                    preview_rows,
                    write_preview,
                )
            )
        temporary_check = prepare_check_workbook(
            summary, month_detail, check_output_path
        )
        commit_prepared_outputs(
            panel_updates, temporary_check, check_output_path
        )
    except Exception:
        for update in panel_updates:
            update.temporary_panel_path.unlink(missing_ok=True)
            if update.temporary_preview_path is not None:
                update.temporary_preview_path.unlink(missing_ok=True)
        if temporary_check is not None:
            temporary_check.unlink(missing_ok=True)
        raise

    return {
        "max_forward": max_forward,
        "added_column_count": len(output_columns.columns),
        "brute_force_checked_rows": brute_force_rows,
        "identity_coverage": identity_coverage,
        "panel_updates": panel_updates,
        "check_output_path": check_output_path,
        "summary": summary,
    }


def print_summary(result: dict[str, object]) -> None:
    """打印运行参数、验证覆盖、写盘结果与 summary 表完整内容。"""
    print(f"最大前看深度：{result['max_forward']} 个月")
    print(f"新增列数：{result['added_column_count']}")
    print(f"暴力复算通过行数：{result['brute_force_checked_rows']}")
    print("恒等式检查覆盖行数：")
    for regime, row_count in result["identity_coverage"].items():
        print(f"  {regime}: {row_count:,}")
    print(f"  合计: {sum(result['identity_coverage'].values()):,}")
    for update in result["panel_updates"]:
        print(f"面板：{update.panel_path}")
        if update.preview_path is not None:
            print(f"  预览：{update.preview_path}")
        if update.replaced_existing_columns:
            print(
                "  重复运行，已替换本脚本旧列数："
                f"{update.replaced_existing_columns}"
            )
        print(f"  行数：{update.row_count:,}；列数：{update.column_count:,}")
    print(f"校验表：{result['check_output_path']}")
    print("summary sheet：")
    print(result["summary"].to_string(index=False))


def main() -> None:
    """命令行入口。"""
    args = parse_args()
    result = generate_market_condition_future_returns(
        source_panel_path=args.source_panel,
        target_paths=args.targets,
        max_forward=args.max_forward,
        check_output_path=args.check_output,
        preview_rows=args.preview_rows,
        write_preview=not args.no_preview,
    )
    print_summary(result)


if __name__ == "__main__":
    main()
